import base64
from io import BytesIO
from calendar import monthrange
from datetime import date, datetime

from openpyxl import load_workbook

from odoo import fields, models
from odoo.exceptions import ValidationError


class HrAttendanceRegisterImportWizard(models.TransientModel):
    _name = 'hr.attendance.register.import.wizard'
    _description = 'Attendance Register Import Wizard'

    file_name = fields.Char(string='File Name')
    file_data = fields.Binary(string='Excel File', required=True)
    result_html = fields.Html(string='Import Result', readonly=True)

    MONTH_NAME_MAP = {
        'jan': 1,
        'january': 1,
        'feb': 2,
        'february': 2,
        'mar': 3,
        'march': 3,
        'apr': 4,
        'april': 4,
        'may': 5,
        'jun': 6,
        'june': 6,
        'jul': 7,
        'july': 7,
        'aug': 8,
        'august': 8,
        'sep': 9,
        'sept': 9,
        'september': 9,
        'oct': 10,
        'october': 10,
        'nov': 11,
        'november': 11,
        'dec': 12,
        'december': 12,
    }

    def _clean_text(self, value):
        if value in (None, False):
            return ''
        return ' '.join(str(value).replace('\n', ' ').strip().split())

    def _normalize_code(self, raw_value):
        if raw_value is None:
            return False

        value = str(raw_value).strip().upper()
        if not value or value == '-':
            return False

        aliases = {
            'P': 'P',
            'PRESENT': 'P',

            'R': 'R',
            'WFH': 'R',
            'REMOTE': 'R',
            'WORK FROM HOME': 'R',

            'H': 'H',
            'HOLIDAY': 'H',

            'S': 'S',
            'SL': 'S',
            'SICK': 'S',
            'SICK LEAVE': 'S',

            'C': 'C',
            'CL': 'C',
            'CASUAL': 'C',
            'CASUAL LEAVE': 'C',

            'U': 'U',
            'UL': 'U',
            'UNPAID': 'U',
            'UNPAID LEAVE': 'U',

            'D': 'D',
            'LATE': 'D',
            'DATA LATE': 'D',

            'OT': 'OT',
            'OVERTIME': 'OT',
        }
        return aliases.get(value, False)

    def _is_date_like(self, value):
        return isinstance(value, (date, datetime))

    def _extract_month_date(self, value, default_year=None):
        """
        Accepts:
        - real Excel dates
        - April
        - April 2026
        - Apr-26
        - 2026-04-01
        """
        if isinstance(value, datetime):
            return value.date().replace(day=1)

        if isinstance(value, date):
            return value.replace(day=1)

        text = self._clean_text(value)
        if not text:
            return False

        for fmt in (
            '%B %Y',
            '%b %Y',
            '%B-%Y',
            '%b-%Y',
            '%B-%y',
            '%b-%y',
            '%Y-%m-%d',
            '%d-%b-%Y',
            '%d-%b-%y',
            '%d/%m/%Y',
            '%m/%d/%Y',
        ):
            try:
                parsed = datetime.strptime(text, fmt).date()
                return parsed.replace(day=1)
            except Exception:
                continue

        lower_text = text.lower()
        if lower_text in self.MONTH_NAME_MAP:
            year_value = default_year or fields.Date.context_today(self).year
            return date(year_value, self.MONTH_NAME_MAP[lower_text], 1)

        return False

    def _get_fiscal_year_label(self, month_date):
        if month_date.month >= 7:
            fy_start = month_date.year
            fy_end = month_date.year + 1
        else:
            fy_start = month_date.year - 1
            fy_end = month_date.year

        return f"FY-{str(fy_start)[-2:]}/{str(fy_end)[-2:]}"

    def _find_employee(self, employee_env, employee_name_text):
        employee_name_text = self._clean_text(employee_name_text)
        if not employee_name_text:
            return False

        employee = employee_env.search([
            ('name', '=', employee_name_text)
        ], limit=1)

        if employee:
            return employee

        # Safer fallback for minor case/spacing differences.
        employee = employee_env.search([
            ('name', '=ilike', employee_name_text)
        ], limit=1)

        return employee or False

    def _looks_like_standard_sheet(self, sheet):
        """
        Old supported format:
        A = Fiscal Year
        B = Month/date
        C = Employee Name
        D onward = daily codes
        """
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30), values_only=True):
            if not row or len(row) < 4:
                continue

            month_cell = row[1] if len(row) > 1 else None
            employee_name = row[2] if len(row) > 2 else None

            if self._extract_month_date(month_cell) and employee_name and not self._is_date_like(employee_name):
                return True

        return False

    def _get_compact_day_columns(self, sheet):
        """
        New compact format:
        Row 1: blank, 1, 2, 3, 4...
        Row 2: blank, Wed, Thu...
        Row 3: April, P, P, P...
        Row 4 onward: Employee Name, daily codes...
        """
        row_1 = [cell.value for cell in sheet[1]]
        day_columns = []

        for zero_index, value in enumerate(row_1):
            if zero_index == 0:
                continue

            if value in (None, ''):
                if day_columns:
                    break
                continue

            try:
                day_no = int(float(value))
            except (TypeError, ValueError):
                if day_columns:
                    break
                continue

            if 1 <= day_no <= 31:
                day_columns.append((day_no, zero_index))
            elif day_columns:
                break

        return day_columns

    def _looks_like_compact_sheet(self, sheet):
        day_columns = self._get_compact_day_columns(sheet)
        if len(day_columns) < 20:
            return False

        month_cell = sheet.cell(row=3, column=1).value
        month_date = self._extract_month_date(
            month_cell,
            default_year=fields.Date.context_today(self).year,
        )

        return True if month_date else False

    def _select_attendance_sheet(self, workbook):
        """
        Prefer active sheet if valid.
        If active sheet is wrong, scan other sheets.
        This helps when user uploads a workbook with attendance and payroll sheets.
        """
        active_sheet = workbook.active

        if self._looks_like_compact_sheet(active_sheet):
            return active_sheet, 'compact'

        if self._looks_like_standard_sheet(active_sheet):
            return active_sheet, 'standard'

        for sheet in workbook.worksheets:
            if self._looks_like_compact_sheet(sheet):
                return sheet, 'compact'

        for sheet in workbook.worksheets:
            if self._looks_like_standard_sheet(sheet):
                return sheet, 'standard'

        raise ValidationError(
            "No supported attendance sheet found. Supported formats: "
            "old format with Fiscal Year/Month/Employee columns, or compact format with day numbers in row 1 and month in A3."
        )

    def _upsert_attendance_code(
        self,
        register_env,
        employee,
        attendance_date,
        attendance_code,
        fiscal_year_label,
        month_label,
    ):
        return register_env.create_or_update_attendance_line(
            employee=employee,
            attendance_date=attendance_date,
            attendance_code=attendance_code,
            fiscal_year_label=fiscal_year_label,
            month_label=month_label,
            source='sheet_import',
            notes=f"Imported from file {self.file_name or ''}",
        )

    def _import_standard_format(self, sheet, employee_env, register_env):
        created_count = 0
        updated_count = 0
        skipped_count = 0
        unknown_employee_count = 0
        unknown_code_count = 0
        details = []

        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
            if not row or not any(row):
                continue

            fiscal_year_cell = row[0] if len(row) > 0 else None
            month_cell = row[1] if len(row) > 1 else None
            employee_name = row[2] if len(row) > 2 else None

            if row_idx == 1 and fiscal_year_cell and not month_cell and not employee_name:
                continue

            if not employee_name or self._is_date_like(employee_name):
                continue

            month_date = self._extract_month_date(month_cell)
            if not month_date:
                skipped_count += 1
                details.append(f"Row {row_idx} skipped: invalid month value.")
                continue

            current_year_number = month_date.year
            current_month_number = month_date.month
            fiscal_year_label = self._clean_text(fiscal_year_cell) or self._get_fiscal_year_label(month_date)
            month_label = month_date.strftime('%B %Y')

            employee_name_text = self._clean_text(employee_name)
            employee = self._find_employee(employee_env, employee_name_text)

            if not employee:
                unknown_employee_count += 1
                skipped_count += 1
                details.append(f"Row {row_idx} skipped: employee '{employee_name_text}' not found.")
                continue

            max_day = monthrange(current_year_number, current_month_number)[1]

            for day_no in range(1, max_day + 1):
                col_index = 3 + (day_no - 1)
                if col_index >= len(row):
                    continue

                raw_code = row[col_index]
                attendance_code = self._normalize_code(raw_code)

                if raw_code not in (None, '', '-') and not attendance_code:
                    unknown_code_count += 1
                    details.append(
                        f"Row {row_idx}, day {day_no}: unknown code '{raw_code}' for {employee.name}."
                    )
                    continue

                if not attendance_code:
                    continue

                attendance_date = date(current_year_number, current_month_number, day_no)

                _line, status = self._upsert_attendance_code(
                    register_env=register_env,
                    employee=employee,
                    attendance_date=attendance_date,
                    attendance_code=attendance_code,
                    fiscal_year_label=fiscal_year_label,
                    month_label=month_label,
                )

                if status == 'created':
                    created_count += 1
                else:
                    updated_count += 1

        return created_count, updated_count, skipped_count, unknown_employee_count, unknown_code_count, details

    def _import_compact_format(self, sheet, employee_env, register_env):
        created_count = 0
        updated_count = 0
        skipped_count = 0
        unknown_employee_count = 0
        unknown_code_count = 0
        details = []

        default_year = fields.Date.context_today(self).year
        month_cell = sheet.cell(row=3, column=1).value
        month_date = self._extract_month_date(month_cell, default_year=default_year)

        if not month_date:
            raise ValidationError("Compact attendance sheet detected, but month value in A3 is invalid.")

        current_year_number = month_date.year
        current_month_number = month_date.month
        max_day = monthrange(current_year_number, current_month_number)[1]
        month_label = month_date.strftime('%B %Y')
        fiscal_year_label = self._get_fiscal_year_label(month_date)

        day_columns = self._get_compact_day_columns(sheet)
        if not day_columns:
            raise ValidationError("Compact attendance sheet detected, but day number columns were not found in row 1.")

        details.append(
            f"Compact attendance format detected on sheet '{sheet.title}'. "
            f"Month imported as {month_label}. Fiscal Year: {fiscal_year_label}."
        )

        # Row 1 = day numbers
        # Row 2 = weekdays
        # Row 3 = month/template row
        # Row 4 onward = employees
        for row_idx in range(4, sheet.max_row + 1):
            employee_name_raw = sheet.cell(row=row_idx, column=1).value
            employee_name_text = self._clean_text(employee_name_raw)

            if not employee_name_text:
                continue

            if employee_name_text.lower() in self.MONTH_NAME_MAP:
                continue

            if employee_name_text.lower() in ('total', 'grand total'):
                continue

            employee = self._find_employee(employee_env, employee_name_text)

            if not employee:
                unknown_employee_count += 1
                skipped_count += 1
                details.append(f"Row {row_idx} skipped: employee '{employee_name_text}' not found.")
                continue

            for day_no, zero_col_index in day_columns:
                if day_no > max_day:
                    continue

                raw_code = sheet.cell(row=row_idx, column=zero_col_index + 1).value
                attendance_code = self._normalize_code(raw_code)

                if raw_code not in (None, '', '-') and not attendance_code:
                    unknown_code_count += 1
                    details.append(
                        f"Row {row_idx}, day {day_no}: unknown code '{raw_code}' for {employee.name}."
                    )
                    continue

                if not attendance_code:
                    continue

                attendance_date = date(current_year_number, current_month_number, day_no)

                _line, status = self._upsert_attendance_code(
                    register_env=register_env,
                    employee=employee,
                    attendance_date=attendance_date,
                    attendance_code=attendance_code,
                    fiscal_year_label=fiscal_year_label,
                    month_label=month_label,
                )

                if status == 'created':
                    created_count += 1
                else:
                    updated_count += 1

        return created_count, updated_count, skipped_count, unknown_employee_count, unknown_code_count, details

    def action_import_attendance_register(self):
        self.ensure_one()

        if not self.file_data:
            raise ValidationError("Please upload an Excel file.")

        file_content = base64.b64decode(self.file_data)
        workbook = load_workbook(BytesIO(file_content), data_only=True)

        sheet, layout = self._select_attendance_sheet(workbook)

        employee_env = self.env['hr.employee'].sudo()
        register_env = self.env['hr.attendance.register.line'].sudo()

        if layout == 'compact':
            (
                created_count,
                updated_count,
                skipped_count,
                unknown_employee_count,
                unknown_code_count,
                details,
            ) = self._import_compact_format(sheet, employee_env, register_env)
        else:
            (
                created_count,
                updated_count,
                skipped_count,
                unknown_employee_count,
                unknown_code_count,
                details,
            ) = self._import_standard_format(sheet, employee_env, register_env)

        self.result_html = f"""
            <h3>Attendance Register Import Summary</h3>
            <ul>
                <li><strong>Detected Sheet:</strong> {sheet.title}</li>
                <li><strong>Detected Format:</strong> {layout}</li>
                <li><strong>Created Lines:</strong> {created_count}</li>
                <li><strong>Updated Lines:</strong> {updated_count}</li>
                <li><strong>Skipped Rows / Items:</strong> {skipped_count}</li>
                <li><strong>Unknown Employees:</strong> {unknown_employee_count}</li>
                <li><strong>Unknown Codes:</strong> {unknown_code_count}</li>
            </ul>
            <h4>Details</h4>
            <ul>
                {''.join(f'<li>{item}</li>' for item in details[:150])}
            </ul>
        """

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'hr.attendance.register.import.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }