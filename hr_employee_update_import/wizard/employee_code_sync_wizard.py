import base64
from io import BytesIO

from odoo import models, fields, _
from odoo.exceptions import UserError

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class HrEmployeeCodeSyncWizard(models.TransientModel):
    _name = 'hr.employee.code.sync.wizard'
    _description = 'HR Employee Code Sync Wizard'

    file_data = fields.Binary(string='Excel File', required=True)
    file_name = fields.Char(string='Filename')
    result_html = fields.Html(string='Sync Result')

    def _normalize_header(self, value):
        return (value or '').strip().lower()

    def _build_row_dict(self, headers, row_values):
        row_dict = {}
        for index, header in enumerate(headers):
            row_dict[self._normalize_header(header)] = row_values[index] if index < len(row_values) else False
        return row_dict

    def action_sync_employee_codes(self):
        self.ensure_one()

        if load_workbook is None:
            raise UserError(_("The Python package 'openpyxl' is not installed on this server."))

        if not self.file_data:
            raise UserError(_("Please upload an Excel file first."))

        file_content = base64.b64decode(self.file_data)
        workbook = load_workbook(filename=BytesIO(file_content), data_only=True)
        sheet = workbook.active

        if sheet.max_row < 2:
            raise UserError(_("The uploaded Excel file does not contain data rows."))

        headers = [cell.value for cell in sheet[1]]
        normalized_headers = [self._normalize_header(header) for header in headers]

        required_headers = ['employee name', 'new employee code']
        missing_headers = [header for header in required_headers if header not in normalized_headers]
        if missing_headers:
            raise UserError(_("Missing required Excel column(s): %s") % ', '.join(missing_headers))

        updated_count = 0
        skipped_count = 0
        error_logs = []

        employee_model = self.env['hr.employee'].sudo()

        for row_number in range(2, sheet.max_row + 1):
            row_values = [sheet.cell(row=row_number, column=col_number).value for col_number in range(1, len(headers) + 1)]

            if not any(row_values):
                continue

            row_dict = self._build_row_dict(headers, row_values)

            employee_name = (row_dict.get('employee name') or '')
            new_employee_code = (row_dict.get('new employee code') or '')

            employee_name = str(employee_name).strip()
            new_employee_code = str(new_employee_code).strip()

            if not employee_name:
                skipped_count += 1
                error_logs.append("Row %s skipped: Employee Name is empty." % row_number)
                continue

            if not new_employee_code:
                skipped_count += 1
                error_logs.append("Row %s skipped: New Employee Code is empty." % row_number)
                continue

            matched_employees = employee_model.search([('name', '=', employee_name)])

            if not matched_employees:
                skipped_count += 1
                error_logs.append("Row %s skipped: Employee '%s' not found." % (row_number, employee_name))
                continue

            if len(matched_employees) > 1:
                skipped_count += 1
                error_logs.append("Row %s skipped: Multiple employees found with name '%s'." % (row_number, employee_name))
                continue

            employee = matched_employees[0]

            existing_code_employee = employee_model.search([
                ('employee_code', '=', new_employee_code),
                ('id', '!=', employee.id),
            ], limit=1)

            if existing_code_employee:
                skipped_count += 1
                error_logs.append(
                    "Row %s skipped: Code '%s' already belongs to '%s'."
                    % (row_number, new_employee_code, existing_code_employee.name)
                )
                continue

            if employee.employee_code == new_employee_code:
                continue

            try:
                employee.write({'employee_code': new_employee_code})
                updated_count += 1
            except Exception as error:
                skipped_count += 1
                error_logs.append("Row %s failed for '%s': %s" % (row_number, employee_name, str(error)))

        result_lines = [
            "<h3>Employee Code Sync Summary</h3>",
            "<ul>",
            "<li><strong>Updated Employees:</strong> %s</li>" % updated_count,
            "<li><strong>Skipped Rows:</strong> %s</li>" % skipped_count,
            "</ul>",
        ]

        if error_logs:
            result_lines.append("<h4>Details</h4><ul>")
            for line in error_logs[:100]:
                result_lines.append("<li>%s</li>" % line)
            result_lines.append("</ul>")

        self.result_html = ''.join(result_lines)

        return {
            'type': 'ir.actions.act_window',
            'name': _('Employee Code Sync'),
            'res_model': 'hr.employee.code.sync.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }