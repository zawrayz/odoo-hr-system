import base64
from io import BytesIO
from datetime import datetime, date

from odoo import models, fields, _
from odoo.exceptions import UserError

try:
    from openpyxl import load_workbook
    from openpyxl.utils.datetime import from_excel
except ImportError:
    load_workbook = None
    from_excel = None


class HrEmployeeUpdateImportWizard(models.TransientModel):
    _name = 'hr.employee.update.import.wizard'
    _description = 'HR Employee Update Import Wizard'

    file_data = fields.Binary(string='Excel File', required=True)
    file_name = fields.Char(string='Filename')
    result_html = fields.Html(string='Import Result', readonly=True)

    def _normalize_header(self, value):
        return (value or '').strip().lower()

    def _parse_date_value(self, value, field_label):
        if not value:
            return False

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        if isinstance(value, (int, float)):
            if from_excel is None:
                raise UserError(_("openpyxl is required to convert Excel serial dates."))
            try:
                excel_date = from_excel(value)
                if isinstance(excel_date, datetime):
                    return excel_date.date()
                return excel_date
            except Exception:
                raise UserError(_("Invalid date value '%s' found in column '%s'.") % (value, field_label))

        if isinstance(value, str):
            text = value.strip()
            if not text:
                return False

            date_formats = [
                '%Y-%m-%d',
                '%d-%m-%Y',
                '%d/%m/%Y',
                '%m/%d/%Y',
                '%d-%b-%y',
                '%d-%b-%Y',
                '%d/%b/%Y',
                '%b %d, %Y',
                '%B %d, %Y',
            ]
            for fmt in date_formats:
                try:
                    return datetime.strptime(text, fmt).date()
                except Exception:
                    continue

            if text.isdigit():
                if from_excel is None:
                    raise UserError(_("openpyxl is required to convert Excel serial dates."))
                try:
                    excel_date = from_excel(int(text))
                    if isinstance(excel_date, datetime):
                        return excel_date.date()
                    return excel_date
                except Exception:
                    pass

        raise UserError(_("Invalid date value '%s' found in column '%s'.") % (value, field_label))

    def _get_cell_value(self, row_dict, header_name):
        return row_dict.get(self._normalize_header(header_name))

    def _build_row_dict(self, headers, row_values):
        row_dict = {}
        for index, header in enumerate(headers):
            row_dict[self._normalize_header(header)] = row_values[index] if index < len(row_values) else False
        return row_dict

    def action_import_employees(self):
        self.ensure_one()

        if load_workbook is None:
            raise UserError(_("The Python package 'openpyxl' is not installed on this server."))

        if not self.file_data:
            raise UserError(_("Please upload an Excel file first."))

        file_content = base64.b64decode(self.file_data)
        workbook = load_workbook(filename=BytesIO(file_content), data_only=True)
        sheet = workbook.active

        if sheet.max_row < 2:
            raise UserError(_("The uploaded Excel file does not contain employee data rows."))

        headers = [cell.value for cell in sheet[1]]
        normalized_headers = [self._normalize_header(header) for header in headers]

        required_headers = [
            'employee code',
            'employee name',
        ]
        missing_headers = [header for header in required_headers if header not in normalized_headers]
        if missing_headers:
            raise UserError(
                _("Missing required Excel column(s): %s") % ', '.join(missing_headers)
            )

        updated_count = 0
        skipped_count = 0
        empty_code_count = 0
        unchanged_count = 0
        error_logs = []

        employee_model = self.env['hr.employee'].sudo()

        for row_number in range(2, sheet.max_row + 1):
            row_values = [sheet.cell(row=row_number, column=col_number).value for col_number in range(1, len(headers) + 1)]

            if not any(row_values):
                continue

            row_dict = self._build_row_dict(headers, row_values)

            employee_code = self._get_cell_value(row_dict, 'Employee Code')
            employee_name = self._get_cell_value(row_dict, 'Employee Name')

            employee_code = (str(employee_code).strip() if employee_code else '')
            employee_name = (str(employee_name).strip() if employee_name else '')

            if not employee_code:
                empty_code_count += 1
                skipped_count += 1
                error_logs.append("Row %s skipped: Employee Code is empty." % row_number)
                continue

            employee = employee_model.search([('employee_code', '=', employee_code)], limit=1)
            if not employee:
                skipped_count += 1
                error_logs.append("Row %s skipped: Employee with code '%s' not found." % (row_number, employee_code))
                continue

            vals = {}

            # Safe text fields: update only if Excel cell is not blank
            text_field_map = {
    'Employee Name': 'name',
    'Father Name': 'father_name',
    'Account Title': 'bank_account_title',
    'Bank Name': 'bank_name_custom',
    'Account Number': 'bank_account_number',
    'Personal Email': 'personal_email',
    'Work Email': 'work_email',
    'CNIC Number': 'cnic_number',
    'Address': 'street_address',
    'Emergency Contact Name': 'emergency_contact_name',
    'Emergency Contact Number': 'emergency_contact_number',
    'Emergency Contact Relation': 'emergency_contact_relation',
            }

            for excel_label, field_name in text_field_map.items():
                value = self._get_cell_value(row_dict, excel_label)
                if value not in (False, None, ''):
                    vals[field_name] = str(value).strip()

            job_position_value = self._get_cell_value(row_dict, 'Job Position')
            if job_position_value not in (False, None, ''):
                job_position_name = str(job_position_value).strip()
                job_record = self.env['hr.job'].sudo().search([('name', '=', job_position_name)], limit=1)
                if job_record:
                    vals['job_id'] = job_record.id            

            # Safe date fields
            date_field_map = {
                'Joining Date': 'joining_date',
                'Date of Birth': 'date_of_birth_custom',
            }

            try:
                for excel_label, field_name in date_field_map.items():
                    value = self._get_cell_value(row_dict, excel_label)
                    if value not in (False, None, ''):
                        vals[field_name] = self._parse_date_value(value, excel_label)
            except UserError as error:
                skipped_count += 1
                error_logs.append("Row %s skipped: %s" % (row_number, error.args[0]))
                continue

            if not vals:
                unchanged_count += 1
                continue

            try:
                employee.write(vals)
                updated_count += 1
            except Exception as error:
                skipped_count += 1
                error_logs.append("Row %s failed for '%s' (%s): %s" % (
                    row_number,
                    employee_name or employee.name or employee_code,
                    employee_code,
                    str(error)
                ))

        result_lines = [
            "<h3>Employee Update Import Summary</h3>",
            "<ul>",
            "<li><strong>Updated Employees:</strong> %s</li>" % updated_count,
            "<li><strong>Skipped Rows:</strong> %s</li>" % skipped_count,
            "<li><strong>Rows with Empty Employee Code:</strong> %s</li>" % empty_code_count,
            "<li><strong>Unchanged Rows:</strong> %s</li>" % unchanged_count,
            "</ul>",
        ]

        if error_logs:
            result_lines.append("<h4>Details</h4><ul>")
            for line in error_logs[:100]:
                result_lines.append("<li>%s</li>" % line)
            result_lines.append("</ul>")

        self.result_html = ''.join(result_lines)

        return {
            'type': 'ir.actions.act_window',
            'name': _('Employee Update Import'),
            'res_model': 'hr.employee.update.import.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }