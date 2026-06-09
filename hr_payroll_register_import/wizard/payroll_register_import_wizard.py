import base64
from io import BytesIO
from datetime import datetime, date

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from odoo import fields, models
from odoo.exceptions import ValidationError


class HrPayrollRegisterImportWizard(models.TransientModel):
    _name = 'hr.payroll.register.import.wizard'
    _description = 'Payroll Register Import Wizard'

    file_name = fields.Char(string='File Name')
    file_data = fields.Binary(string='Excel File', required=True)
    result_html = fields.Html(string='Import Result', readonly=True)

    MONTH_NAME_MAP = {
        'jan': 1, 'january': 1,
        'feb': 2, 'february': 2,
        'mar': 3, 'march': 3,
        'apr': 4, 'april': 4,
        'may': 5,
        'jun': 6, 'june': 6,
        'jul': 7, 'july': 7,
        'aug': 8, 'august': 8,
        'sep': 9, 'sept': 9, 'september': 9,
        'oct': 10, 'october': 10,
        'nov': 11, 'november': 11,
        'dec': 12, 'december': 12,
    }

    def _clean_text(self, value):
        if value in (None, False):
            return ''
        return ' '.join(str(value).replace('\n', ' ').strip().split())

    def _to_float(self, value):
        if value in (None, '', False):
            return 0.0

        try:
            clean_value = str(value).replace(',', '').strip()
            if not clean_value:
                return 0.0
            return float(clean_value)
        except Exception:
            return 0.0

    def _to_date(self, value):
        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        if isinstance(value, (int, float)) and 20000 <= value <= 70000:
            try:
                excel_date = from_excel(value)
                if isinstance(excel_date, datetime):
                    return excel_date.date()
                if isinstance(excel_date, date):
                    return excel_date
            except Exception:
                return False

        text = self._clean_text(value)
        if not text:
            return False

        for fmt in (
            '%Y-%m-%d',
            '%d-%b-%Y',
            '%d-%b-%y',
            '%d/%m/%Y',
            '%m/%d/%Y',
            '%B %Y',
            '%b %Y',
            '%B-%Y',
            '%b-%Y',
            '%B-%y',
            '%b-%y',
        ):
            try:
                parsed = datetime.strptime(text, fmt).date()
                return parsed
            except Exception:
                continue

        lower_text = text.lower()
        if lower_text in self.MONTH_NAME_MAP:
            today = fields.Date.context_today(self)
            return date(today.year, self.MONTH_NAME_MAP[lower_text], 1)

        return False

    def _month_start(self, value):
        parsed_date = self._to_date(value)
        if parsed_date:
            return parsed_date.replace(day=1)
        return False

    def _get_fiscal_year_label(self, month_date):
        if not month_date:
            return ''

        if month_date.month >= 7:
            fy_start = month_date.year
            fy_end = month_date.year + 1
        else:
            fy_start = month_date.year - 1
            fy_end = month_date.year

        return f"FY-{str(fy_start)[-2:]}/{str(fy_end)[-2:]}"

    def _normalize_fiscal_year_text(self, value, month_date=False):
        text = self._clean_text(value)

        if text:
            cleaned = text.replace(' ', '')
            if cleaned.upper().startswith('FY'):
                cleaned = cleaned[2:]
            cleaned = cleaned.replace('-', '/')

            if '/' in cleaned:
                parts = cleaned.split('/')
                if len(parts) == 2:
                    try:
                        start = int(parts[0])
                        end = int(parts[1])
                        if start < 100 and end < 100:
                            return f"FY-{start:02d}/{end:02d}"
                    except Exception:
                        pass

        return self._get_fiscal_year_label(month_date)

    def _build_header_map(self, sheet, header_row):
        """
        Supports:
        - single-row headers
        - two-row headers where row above contains Payment Date / Total Salary / Loans from Cash
        """
        header_map = {}

        current_values = [
            self._clean_text(sheet.cell(row=header_row, column=col).value)
            for col in range(1, sheet.max_column + 1)
        ]

        previous_values = []
        if header_row > 1:
            previous_values = [
                self._clean_text(sheet.cell(row=header_row - 1, column=col).value)
                for col in range(1, sheet.max_column + 1)
            ]
        else:
            previous_values = [''] * sheet.max_column

        for idx in range(sheet.max_column):
            current = current_values[idx] if idx < len(current_values) else ''
            previous = previous_values[idx] if idx < len(previous_values) else ''

            header = current or previous

            if not header:
                continue

            # Keep first duplicate. Example: "Total" appears twice in some sheets.
            if header not in header_map:
                header_map[header] = idx

            if previous and current:
                combined = f"{previous} / {current}"
                if combined not in header_map:
                    header_map[combined] = idx

        return header_map

    def _get_col(self, header_map, *names):
        for name in names:
            if name in header_map:
                return header_map[name]

        lowered_map = {
            self._clean_text(key).lower(): value
            for key, value in header_map.items()
        }

        for name in names:
            lowered_name = self._clean_text(name).lower()
            if lowered_name in lowered_map:
                return lowered_map[lowered_name]

        return False

    def _row_has_payroll_header(self, sheet, row_number):
        values = [
            self._clean_text(sheet.cell(row=row_number, column=col).value)
            for col in range(1, sheet.max_column + 1)
        ]
        value_set = set(values)

        return 'Employee ID' in value_set and 'Employee Name' in value_set

    def _find_payroll_header_rows(self, sheet):
        header_rows = []

        for row_number in range(1, sheet.max_row + 1):
            if self._row_has_payroll_header(sheet, row_number):
                header_rows.append(row_number)

        return header_rows

    def _sheet_has_payroll_header(self, sheet):
        return bool(self._find_payroll_header_rows(sheet))

    def _select_payroll_sheet(self, workbook):
        active_sheet = workbook.active

        if self._sheet_has_payroll_header(active_sheet):
            return active_sheet

        for sheet in workbook.worksheets:
            if self._sheet_has_payroll_header(sheet):
                return sheet

        raise ValidationError(
            "Payroll sheet header could not be detected. Please check that Employee ID and Employee Name columns exist."
        )

    def _find_metadata_near_header(self, sheet, header_row):
        month_date = False
        payment_date = False
        fiscal_year_label = ''

        start_row = max(1, header_row - 15)

        for row_number in range(start_row, header_row):
            for col_number in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_number, column=col_number).value
                text = self._clean_text(cell_value)
                text_lower = text.lower()

                if not text:
                    continue

                if 'for the month' in text_lower or text_lower in ('month', 'payroll month'):
                    for offset in range(1, 5):
                        candidate = sheet.cell(row=row_number, column=col_number + offset).value
                        candidate_date = self._month_start(candidate)
                        if candidate_date:
                            month_date = candidate_date
                            break

                if text_lower in ('date of payment', 'payment date'):
                    for offset in range(1, 5):
                        candidate = sheet.cell(row=row_number, column=col_number + offset).value
                        candidate_date = self._to_date(candidate)
                        if candidate_date:
                            payment_date = candidate_date
                            break

                if text.upper().startswith('FY'):
                    fiscal_year_label = self._normalize_fiscal_year_text(text, month_date)

        if month_date and not fiscal_year_label:
            fiscal_year_label = self._get_fiscal_year_label(month_date)

        return month_date, payment_date, fiscal_year_label

    def _get_column_indexes(self, header_map):
        return {
            'month_idx': self._get_col(header_map, 'Month', 'Payroll Month'),
            'fy_idx': self._get_col(header_map, 'FY Selection', 'Fiscal Year', 'FY'),

            'employee_code_idx': self._get_col(header_map, 'Employee ID', 'Employee Code'),
            'employee_name_idx': self._get_col(header_map, 'Employee Name'),

            'bank_idx': self._get_col(header_map, 'Bank'),
            'designation_idx': self._get_col(header_map, 'Designation'),
            'payment_method_idx': self._get_col(header_map, 'Payment method', 'Payment Method'),

            'basic_salary_idx': self._get_col(header_map, 'Basic Salary'),
            'basic_actual_idx': self._get_col(header_map, 'Basic Actual'),
            'medical_allowance_idx': self._get_col(header_map, 'Medical Allowance'),
            'advertised_salary_idx': self._get_col(header_map, 'Advertised Salary'),

            'project_salary_idx': self._get_col(header_map, 'Project Salary'),
            'project_detail_idx': self._get_col(header_map, 'Project Salary / Project', 'Project'),

            'bonus_idx': self._get_col(header_map, 'Bonus'),
            'bonus_detail_idx': self._get_col(header_map, 'Bonus / For', 'For'),

            'allowance_idx': self._get_col(header_map, 'Allowance'),
            'allowance_detail_idx': self._get_col(header_map, 'Allowance Detail'),

            'overtime_idx': self._get_col(header_map, 'Overtime'),
            'overtime_detail_idx': self._get_col(header_map, 'OT Detail'),

            'taxable_income_idx': self._get_col(header_map, 'Taxable Income'),
            'yearly_income_idx': self._get_col(header_map, 'Yearly Income'),
            'income_tax_idx': self._get_col(header_map, 'Income Tax Deduction', 'Income Tax'),
            'other_deduction_idx': self._get_col(header_map, 'Other Deductions'),
            'deduction_detail_idx': self._get_col(header_map, 'Deductions for', 'Deduction For'),

            'total_idx': self._get_col(header_map, 'Total'),
            'total_round_idx': self._get_col(header_map, 'Total Round'),
            'total_salary_idx': self._get_col(header_map, 'Total Salary'),
            'loans_idx': self._get_col(header_map, 'Loans from Cash'),

            'payment_date_idx': self._get_col(header_map, 'Payment Date'),
            'payment_idx': self._get_col(header_map, 'Payment', 'last Payment', 'Last Payment'),
            'inclusive_tax_idx': self._get_col(header_map, 'Inclusive of Tax', 'ALFA enrolled'),
            'remaining_idx': self._get_col(header_map, 'Remaining'),

            'period_idx': self._get_col(header_map, 'For the period of'),
            'paid_at_idx': self._get_col(header_map, 'and payed at', 'Paid At'),
            'comments_idx': self._get_col(header_map, 'Comments'),
        }

    def _get_row_value(self, row_values, index, default=''):
        if index is False:
            return default

        if index >= len(row_values):
            return default

        return row_values[index]

    def _find_employee(self, employee_env, employee_code, employee_name):
        employee_code = self._clean_text(employee_code)
        employee_name = self._clean_text(employee_name)

        employee = False

        if employee_code:
            employee = employee_env.search([
                ('employee_code', '=', employee_code)
            ], limit=1)

        if not employee and employee_name:
            employee = employee_env.search([
                ('name', '=', employee_name)
            ], limit=1)

        if not employee and employee_name:
            employee = employee_env.search([
                ('name', '=ilike', employee_name)
            ], limit=1)

        return employee or False

    def _import_payroll_block(
        self,
        sheet,
        header_row,
        next_header_row,
        employee_env,
        payroll_env,
    ):
        created_count = 0
        updated_count = 0
        skipped_count = 0
        unknown_employee_count = 0
        details = []

        header_map = self._build_header_map(sheet, header_row)
        cols = self._get_column_indexes(header_map)

        required_keys = ['employee_code_idx', 'employee_name_idx']
        if any(cols.get(key) is False for key in required_keys):
            skipped_count += 1
            details.append(f"Header row {header_row} skipped: Employee ID / Employee Name not found.")
            return created_count, updated_count, skipped_count, unknown_employee_count, details

        block_month_date, block_payment_date, block_fiscal_year_label = self._find_metadata_near_header(
            sheet,
            header_row,
        )

        if not block_month_date and cols.get('month_idx') is False:
            skipped_count += 1
            details.append(
                f"Header row {header_row} skipped: payroll month not found near header and no Month column exists."
            )
            return created_count, updated_count, skipped_count, unknown_employee_count, details

        data_start_row = header_row + 1
        data_end_row = (next_header_row - 1) if next_header_row else sheet.max_row

        details.append(
            f"Payroll header detected on row {header_row}. "
            f"Month: {block_month_date.strftime('%B %Y') if block_month_date else 'from row column'}."
        )

        for row_idx in range(data_start_row, data_end_row + 1):
            row_values = [
                sheet.cell(row=row_idx, column=col).value
                for col in range(1, sheet.max_column + 1)
            ]

            if not any(row_values):
                continue

            employee_code = self._clean_text(
                self._get_row_value(row_values, cols['employee_code_idx'])
            )
            employee_name = self._clean_text(
                self._get_row_value(row_values, cols['employee_name_idx'])
            )

            if not employee_code and not employee_name:
                continue

            # Skip accidental non-data rows.
            if employee_code.lower() in ('employee id', '#') or employee_name.lower() == 'employee name':
                continue

            row_month_date = False
            if cols.get('month_idx') is not False:
                row_month_date = self._month_start(
                    self._get_row_value(row_values, cols['month_idx'])
                )

            month_date = row_month_date or block_month_date
            if not month_date:
                skipped_count += 1
                details.append(f"Row {row_idx} skipped: missing payroll month.")
                continue

            row_fiscal_year = ''
            if cols.get('fy_idx') is not False:
                row_fiscal_year = self._clean_text(
                    self._get_row_value(row_values, cols['fy_idx'])
                )

            fiscal_year_label = (
                self._normalize_fiscal_year_text(row_fiscal_year, month_date)
                or block_fiscal_year_label
                or self._get_fiscal_year_label(month_date)
            )

            employee = self._find_employee(
                employee_env=employee_env,
                employee_code=employee_code,
                employee_name=employee_name,
            )

            if not employee:
                unknown_employee_count += 1
                skipped_count += 1
                details.append(
                    f"Row {row_idx} skipped: employee not found for code '{employee_code}' / name '{employee_name}'."
                )
                continue

            payment_date = self._to_date(
                self._get_row_value(row_values, cols['payment_date_idx'])
            ) or block_payment_date

            vals = {
                'employee_name_text': employee_name,
                'fiscal_year_label': fiscal_year_label,

                'designation': self._clean_text(self._get_row_value(row_values, cols['designation_idx'])),
                'payment_method': self._clean_text(self._get_row_value(row_values, cols['payment_method_idx'])),

                'basic_salary': self._to_float(self._get_row_value(row_values, cols['basic_salary_idx'], 0)),
                'basic_actual': self._to_float(self._get_row_value(row_values, cols['basic_actual_idx'], 0)),
                'medical_allowance': self._to_float(self._get_row_value(row_values, cols['medical_allowance_idx'], 0)),
                'advertised_salary': self._to_float(self._get_row_value(row_values, cols['advertised_salary_idx'], 0)),

                'project_salary': self._to_float(self._get_row_value(row_values, cols['project_salary_idx'], 0)),
                'project_detail': self._clean_text(self._get_row_value(row_values, cols['project_detail_idx'])),

                'bonus': self._to_float(self._get_row_value(row_values, cols['bonus_idx'], 0)),
                'bonus_detail': self._clean_text(self._get_row_value(row_values, cols['bonus_detail_idx'])),

                'allowance': self._to_float(self._get_row_value(row_values, cols['allowance_idx'], 0)),
                'allowance_detail': self._clean_text(self._get_row_value(row_values, cols['allowance_detail_idx'])),

                'overtime': self._to_float(self._get_row_value(row_values, cols['overtime_idx'], 0)),
                'overtime_detail': self._clean_text(self._get_row_value(row_values, cols['overtime_detail_idx'])),

                'taxable_income': self._to_float(self._get_row_value(row_values, cols['taxable_income_idx'], 0)),
                'yearly_income': self._to_float(self._get_row_value(row_values, cols['yearly_income_idx'], 0)),
                'income_tax_deduction': self._to_float(self._get_row_value(row_values, cols['income_tax_idx'], 0)),
                'other_deductions': self._to_float(self._get_row_value(row_values, cols['other_deduction_idx'], 0)),
                'deduction_detail': self._clean_text(self._get_row_value(row_values, cols['deduction_detail_idx'])),

                'total': self._to_float(self._get_row_value(row_values, cols['total_idx'], 0)),
                'total_round': self._to_float(self._get_row_value(row_values, cols['total_round_idx'], 0)),
                'total_salary': self._to_float(self._get_row_value(row_values, cols['total_salary_idx'], 0)),
                'loans_from_cash': self._to_float(self._get_row_value(row_values, cols['loans_idx'], 0)),

                'payment_date': payment_date,
                'payment': self._to_float(self._get_row_value(row_values, cols['payment_idx'], 0)),
                'inclusive_of_tax': self._clean_text(self._get_row_value(row_values, cols['inclusive_tax_idx'])),
                'remaining': self._to_float(self._get_row_value(row_values, cols['remaining_idx'], 0)),

                'period_label': self._clean_text(self._get_row_value(row_values, cols['period_idx'])),
                'paid_at': self._clean_text(self._get_row_value(row_values, cols['paid_at_idx'])),
                'comments': self._clean_text(self._get_row_value(row_values, cols['comments_idx'])),

                'source': 'sheet_import',
            }

            rec, status = payroll_env.create_or_update_payroll_line(employee, month_date, vals)

            if status == 'created':
                created_count += 1
            else:
                updated_count += 1

        return created_count, updated_count, skipped_count, unknown_employee_count, details

    def action_import_payroll_register(self):
        self.ensure_one()

        if not self.file_data:
            raise ValidationError("Please upload an Excel file.")

        file_content = base64.b64decode(self.file_data)
        workbook = load_workbook(BytesIO(file_content), data_only=True)

        sheet = self._select_payroll_sheet(workbook)

        employee_env = self.env['hr.employee'].sudo()
        payroll_env = self.env['hr.payroll.register.line'].sudo()

        header_rows = self._find_payroll_header_rows(sheet)
        if not header_rows:
            raise ValidationError(
                "Payroll sheet header could not be detected. Please check that Employee ID and Employee Name columns exist."
            )

        total_created = 0
        total_updated = 0
        total_skipped = 0
        total_unknown = 0
        all_details = [
            f"Detected payroll sheet: {sheet.title}.",
            f"Detected payroll header rows: {', '.join(str(row) for row in header_rows)}.",
        ]

        for index, header_row in enumerate(header_rows):
            next_header_row = header_rows[index + 1] if index + 1 < len(header_rows) else False

            (
                created_count,
                updated_count,
                skipped_count,
                unknown_employee_count,
                details,
            ) = self._import_payroll_block(
                sheet=sheet,
                header_row=header_row,
                next_header_row=next_header_row,
                employee_env=employee_env,
                payroll_env=payroll_env,
            )

            total_created += created_count
            total_updated += updated_count
            total_skipped += skipped_count
            total_unknown += unknown_employee_count
            all_details.extend(details)

        self.result_html = f"""
            <h3>Payroll Register Import Summary</h3>
            <ul>
                <li><strong>Detected Sheet:</strong> {sheet.title}</li>
                <li><strong>Created Lines:</strong> {total_created}</li>
                <li><strong>Updated Lines:</strong> {total_updated}</li>
                <li><strong>Skipped Rows:</strong> {total_skipped}</li>
                <li><strong>Unknown Employees:</strong> {total_unknown}</li>
            </ul>
            <h4>Details</h4>
            <ul>
                {''.join(f'<li>{item}</li>' for item in all_details[:150])}
            </ul>
        """

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'hr.payroll.register.import.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }