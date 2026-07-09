import pytz
import time as pytime
from datetime import timedelta,  datetime, time, timedelta, date
import calendar

from odoo import http, fields
from odoo.tools import config
from odoo.exceptions import ValidationError
from odoo.http import request
from urllib.parse import urlencode
from werkzeug.urls import url_encode
from io import BytesIO
from openpyxl import Workbook
import hmac
import hashlib


class HrEmployeePortal(http.Controller):

    SICK_POLICY_DAYS = 10.0
    CASUAL_POLICY_DAYS = 20.0
    SICK_MONTHLY_EARN = round(SICK_POLICY_DAYS / 12.0, 2)
    CASUAL_MONTHLY_EARN = round(CASUAL_POLICY_DAYS / 12.0, 2)

    REQUEST_TYPE_OPTIONS = [
        ('short_leave', 'Request for Short Leave'),
        ('wfh', 'Request for WFH'),
        ('sick_leave', 'Request for Sick Leave'),
        ('casual_leave', 'Request for Casual Leave'),
    ]

    REQUEST_EMAIL_TO = 'noreply@blimpglobal.com'

    ADMIN_PAYROLL_TAX_SLABS = [
        {
            'slab': 'Slab 1',
            'range_label': 'Taxable income does not exceed Rs600,000',
            'tax_rule': '0%',
        },
        {
            'slab': 'Slab 2',
            'range_label': '600,001 to 1,200,000',
            'tax_rule': '1% of the amount exceeding Rs600,000',
        },
        {
            'slab': 'Slab 3',
            'range_label': '1,200,001 to 2,200,000',
            'tax_rule': 'Rs6,000 + 11% of the amount exceeding 1,200,000',
        },
        {
            'slab': 'Slab 4',
            'range_label': '2,200,001 to 3,200,000',
            'tax_rule': 'Rs116,000 + 23% of the amount exceeding Rs2,200,000',
        },
        {
            'slab': 'Slab 5',
            'range_label': '3,200,001 to 4,100,000',
            'tax_rule': 'Rs346,000 + 30% of the amount exceeding Rs3,200,000',
        },
        {
            'slab': 'Slab 6',
            'range_label': 'Taxable income exceeds Rs4,100,000',
            'tax_rule': 'Rs616,000 + 35% of the amount exceeding Rs4,100,000',
        },
    ]

    # ---------------------------------------------------------
    # Helper: Check if user is HR Manager
    # ---------------------------------------------------------
    def _is_hr_manager(self):
        return request.env.user.has_group('hr.group_hr_manager')
    def _get_admin_payroll_register_lines(self, selected_month, employee_id=None):
        if not selected_month:
            return request.env['hr.payroll.register.line'].sudo()

        month_start = selected_month.replace(day=1)
        domain = [('month_date', '=', month_start)]

        if employee_id:
            domain.append(('employee_id', '=', employee_id))

        return request.env['hr.payroll.register.line'].sudo().search(
            domain,
            order='employee_id asc'
        )

    def _get_admin_payroll_data(self, selected_month, employee_id=None):
        payroll_lines = self._get_admin_payroll_register_lines(
            selected_month,
            employee_id=employee_id,
        )

        rows = []
        for line in payroll_lines:
            project_salary_value = self._to_amount(getattr(line, 'project_salary', 0.0))
            allowance_value = self._to_amount(getattr(line, 'allowance', 0.0))
            project_salary_allowance = self._to_amount(project_salary_value + allowance_value)

            project_detail = (
                getattr(line, 'project_detail', False)
                or getattr(line, 'allowance_detail', False)
                or '-'
            )
            bonus_detail = getattr(line, 'bonus_detail', False) or '-'
            deduction_detail = (
                getattr(line, 'deduction_detail', False)
                or getattr(line, 'comments', False)
                or '-'
            )

            rows.append({
                'record_id': line.id,
                'employee_code': line.employee_code or '-',
                'employee_name': line.employee_id.name or line.employee_name_text or '-',
                'bank_name': line.payment_method or '-',
                'designation': line.designation or '-',
                'payment_method': line.payment_method or '-',

                'basic_salary': self._to_amount(line.basic_salary),
                'basic_actual': self._to_amount(line.basic_actual),
                'medical_allowance': self._to_amount(line.medical_allowance),
                'advertised_salary': self._to_amount(line.advertised_salary),

                'project_salary': project_salary_allowance,
                'project_value': project_detail,

                'bonus': self._to_amount(line.bonus),
                'for_value': bonus_detail,

                'overtime': self._to_amount(getattr(line, 'overtime', 0.0)),
                'ot_detail': getattr(line, 'overtime_detail', False) or '-',

                'taxable_income': self._to_amount(line.taxable_income),
                'yearly_income': self._to_amount(line.yearly_income),
                'income_tax_deduction': self._to_amount(line.income_tax_deduction),
                'other_deductions': self._to_amount(line.other_deductions),
                'deduction_for': deduction_detail,

                'total': self._to_amount(line.total),
                'total_round': self._to_amount(line.total_round),
                'total_allowance': allowance_value,
                'hr_cost_including_bonus': self._to_amount(line.total_salary),
                'hr_cost_rounded': self._to_amount(line.total_round),

                'payment_date': line.payment_date.strftime('%d-%b-%Y') if line.payment_date else '-',
                'payment_date_value': line.payment_date.strftime('%Y-%m-%d') if line.payment_date else '',
                'month_label': line.month_label or '-',

                'project_detail_raw': (
                    getattr(line, 'project_detail', False)
                    or getattr(line, 'allowance_detail', False)
                    or ''
                ),
                'bonus_detail_raw': getattr(line, 'bonus_detail', False) or '',
                'deduction_detail_raw': getattr(line, 'deduction_detail', False) or '',
                'comments_raw': line.comments or '',
            })

        totals = {
            'basic_salary': self._sum_row_values(rows, 'basic_salary'),
            'basic_actual': self._sum_row_values(rows, 'basic_actual'),
            'medical_allowance': self._sum_row_values(rows, 'medical_allowance'),
            'advertised_salary': self._sum_row_values(rows, 'advertised_salary'),
            'project_salary': self._sum_row_values(rows, 'project_salary'),
            'project_value': 0.0,
            'bonus': self._sum_row_values(rows, 'bonus'),
            'for_value': 0.0,
            'overtime': self._sum_row_values(rows, 'overtime'),
            'ot_detail': 0.0,
            'taxable_income': self._sum_row_values(rows, 'taxable_income'),
            'yearly_income': self._sum_row_values(rows, 'yearly_income'),
            'income_tax_deduction': self._sum_row_values(rows, 'income_tax_deduction'),
            'other_deductions': self._sum_row_values(rows, 'other_deductions'),
            'deduction_for': 0.0,
            'total': self._sum_row_values(rows, 'total'),
            'total_round': self._sum_row_values(rows, 'total_round'),
            'total_allowance': self._sum_row_values(rows, 'total_allowance'),
            'hr_cost_including_bonus': self._sum_row_values(rows, 'hr_cost_including_bonus'),
            'hr_cost_rounded': self._sum_row_values(rows, 'hr_cost_rounded'),
        }

        calculator_input = rows[0]['total_round'] if rows else 0.0
        calculator_medical = rows[0]['medical_allowance'] if rows else 0.0
        calculator_taxable_yearly = self._to_amount(calculator_input * 12.0)
        calculator_monthly_tax = rows[0]['income_tax_deduction'] if rows else 0.0
        calculator_yearly_tax = self._to_amount(calculator_monthly_tax * 12.0)

        overview = self._get_admin_payroll_overview(selected_month)

        return {
            'overview': overview,
            'fiscal_year_code': overview['fiscal_year_code'],
            'fiscal_year_range_label': overview['fiscal_year_range_label'],
            'rows': rows,
            'totals': totals,
            'tax_slabs': self.ADMIN_PAYROLL_TAX_SLABS,
            'calculator': {
                'put_net_salary': calculator_input,
                'medical_allowance': calculator_medical,
                'taxable_income_yearly': calculator_taxable_yearly,
                'monthly_tax': calculator_monthly_tax,
                'yearly_tax': calculator_yearly_tax,
            },
        }

    # ---------------------------------------------------------
    # Redirect /my safely for employee-linked users only
    # ---------------------------------------------------------
    @http.route('/my', type='http', auth='user', website=True)
    def redirect_my_portal(self, **kwargs):
        if self._is_hr_manager():
            return request.redirect('/my/hr/admin')

        user = request.env.user
        employee = request.env['hr.employee'].sudo().search(
            [('user_id', '=', user.id)],
            limit=1
        )

        if employee:
            return request.redirect('/my/hr')

        return request.render('portal.portal_my_home', {})

    # ---------------------------------------------------------
    # Get logged in employee
    # ---------------------------------------------------------
    def _get_employee(self):
        user = request.env.user
        employee = request.env['hr.employee'].sudo().search(
            [('user_id', '=', user.id)],
            limit=1
        )
        return employee

    # ---------------------------------------------------------
    # Format worked hours (1h 35m)
    # ---------------------------------------------------------
    def _format_hours(self, hours):
        if not hours:
            return "0h"

        total_minutes = int(round(hours * 60))
        hour_value = total_minutes // 60
        minute_value = total_minutes % 60

        if minute_value:
            return f"{hour_value}h {minute_value}m"
        return f"{hour_value}h"

    # ---------------------------------------------------------
    # Format amount safely
    # ---------------------------------------------------------
    def _to_amount(self, value):
        try:
            return round(float(value or 0.0), 2)
        except (TypeError, ValueError):
            return 0.0

    # ---------------------------------------------------------
    # Common page values
    # ---------------------------------------------------------
    def _prepare_portal_values(self, employee=None, extra_values=None):
        values = {
            'employee': employee,
            'is_hr_manager': self._is_hr_manager(),
            'request_type_options': self.REQUEST_TYPE_OPTIONS,
        }
        if extra_values:
            values.update(extra_values)
        return values

    # ---------------------------------------------------------
    # Redirect safely when employee is missing
    # ---------------------------------------------------------
    def _redirect_if_no_employee(self, employee):
        if not employee:
            return request.redirect('/my/hr/profile?missing_employee=1')
        return None

    # ---------------------------------------------------------
    # Parse month from query string (YYYY-MM)
    # ---------------------------------------------------------
    def _get_selected_month_date(self, month_value=None):
        if month_value:
            try:
                parsed = datetime.strptime(month_value, '%Y-%m').date()
                return parsed.replace(day=1)
            except ValueError:
                pass
        today = fields.Date.context_today(request.env.user)
        return today.replace(day=1)

    # ---------------------------------------------------------
    # Previous / Next month values
    # ---------------------------------------------------------
    def _get_month_navigation(self, month_start):
        previous_month_last_day = month_start - timedelta(days=1)
        next_month_first_day = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)

        return {
            'previous_month_value': previous_month_last_day.strftime('%Y-%m'),
            'next_month_value': next_month_first_day.strftime('%Y-%m'),
            'month_display': month_start.strftime('%B %Y'),
        }

    # ---------------------------------------------------------
    # Fiscal year helpers (July -> June)
    # ---------------------------------------------------------
    def _get_fiscal_year_bounds(self, target_date):
        if target_date.month >= 7:
            fy_start = date(target_date.year, 7, 1)
            fy_end = date(target_date.year + 1, 6, 30)
        else:
            fy_start = date(target_date.year - 1, 7, 1)
            fy_end = date(target_date.year, 6, 30)
        return fy_start, fy_end

    def _get_fiscal_year_label(self, target_date):
        fy_start, fy_end = self._get_fiscal_year_bounds(target_date)
        return f"FY-{str(fy_start.year)[-2:]}/{str(fy_end.year)[-2:]}"

    def _get_fiscal_year_code(self, target_date):
        fy_start, fy_end = self._get_fiscal_year_bounds(target_date)
        return f"{str(fy_start.year)[-2:]}/{str(fy_end.year)[-2:]}"

    def _get_fiscal_year_range_label(self, target_date):
        fy_start, fy_end = self._get_fiscal_year_bounds(target_date)
        return f"{fy_start.day} {fy_start.strftime('%B %Y')} to {fy_end.day} {fy_end.strftime('%B %Y')}"

    def _get_fiscal_year_month_starts(self, target_date):
        fy_start, _fy_end = self._get_fiscal_year_bounds(target_date)
        months = []
        current_month = fy_start
        for _i in range(12):
            months.append(current_month)
            current_month = (current_month.replace(day=28) + timedelta(days=4)).replace(day=1)
        return months

    def _get_fiscal_year_month_options(self, target_date):
        month_options = []
        for month_start in self._get_fiscal_year_month_starts(target_date):
            month_options.append({
                'value': month_start.strftime('%Y-%m'),
                'label': month_start.strftime('%B %Y'),
                'short_label': month_start.strftime('%b-%y'),
            })
        return month_options

    def _get_fiscal_year_context(self, target_date):
        fy_start, fy_end = self._get_fiscal_year_bounds(target_date)
        return {
            'fiscal_year_start': fy_start,
            'fiscal_year_end': fy_end,
            'fiscal_year_label': self._get_fiscal_year_label(target_date),
            'fiscal_year_code': self._get_fiscal_year_code(target_date),
            'fiscal_year_range_label': self._get_fiscal_year_range_label(target_date),
            'fiscal_year_month_options': self._get_fiscal_year_month_options(target_date),
        }

    # ---------------------------------------------------------
    # Payroll FY selection helpers
    # ---------------------------------------------------------
    def _get_payroll_fiscal_year_options(self):
        today = fields.Date.context_today(request.env.user)
        current_fy_start, _current_fy_end = self._get_fiscal_year_bounds(today)

        start_years = set()

        # Add old years from existing payroll records
        Payroll = request.env['hr.payroll.register.line'].sudo()
        payroll_months = Payroll.search([('month_date', '!=', False)]).mapped('month_date')

        for month_date in payroll_months:
            fy_start, _fy_end = self._get_fiscal_year_bounds(month_date)
            start_years.add(fy_start.year)

        # Add previous year, current year, and next 5 future years
        for year in range(current_fy_start.year - 1, current_fy_start.year + 6):
            start_years.add(year)

        options = []
        for start_year in sorted(start_years):
            end_year = start_year + 1
            options.append({
                'value': f'{start_year}-{end_year}',
                'label': f'{str(start_year)[-2:]}/{str(end_year)[-2:]}',
                'start_year': start_year,
                'end_year': end_year,
            })

        return options

    def _parse_payroll_fiscal_year_value(self, fy_value):
        if not fy_value:
            return False

        fy_value = str(fy_value).strip()

        try:
            if '-' in fy_value:
                parts = fy_value.split('-')
                if len(parts) == 2:
                    start_year = int(parts[0])
                    end_year = int(parts[1])
                    if end_year == start_year + 1:
                        return date(start_year, 7, 1), date(end_year, 6, 30)

            if '/' in fy_value:
                parts = fy_value.split('/')
                if len(parts) == 2:
                    start_two = int(parts[0])
                    end_two = int(parts[1])

                    if start_two <= 99 and end_two <= 99:
                        start_year = 2000 + start_two
                        end_year = 2000 + end_two

                        if end_year == start_year + 1:
                            return date(start_year, 7, 1), date(end_year, 6, 30)
        except (TypeError, ValueError):
            return False

        return False

    def _get_selected_payroll_period(self, fy_value=None, month_value=None):
        today = fields.Date.context_today(request.env.user)
        today_month = today.replace(day=1)

        parsed_fy = self._parse_payroll_fiscal_year_value(fy_value)
        parsed_month = self._get_selected_month_date(month_value) if month_value else False

        if parsed_fy:
            fy_start, fy_end = parsed_fy

            if parsed_month and fy_start <= parsed_month <= fy_end:
                selected_month = parsed_month
            else:
                if fy_start <= today_month <= fy_end:
                    selected_month = today_month
                else:
                    selected_month = fy_start
        else:
            if parsed_month:
                selected_month = parsed_month
                fy_start, fy_end = self._get_fiscal_year_bounds(selected_month)
            else:
                selected_month = today_month
                fy_start, fy_end = self._get_fiscal_year_bounds(selected_month)

        if selected_month < fy_start or selected_month > fy_end:
            selected_month = fy_start

        selected_fy_value = f'{fy_start.year}-{fy_end.year}'
        selected_fy_code = f'{str(fy_start.year)[-2:]}/{str(fy_end.year)[-2:]}'

        month_options = []
        current_month = fy_start
        for _i in range(12):
            month_options.append({
                'value': current_month.strftime('%Y-%m'),
                'label': current_month.strftime('%B %Y'),
                'short_label': current_month.strftime('%b-%y'),
            })
            current_month = (current_month.replace(day=28) + timedelta(days=4)).replace(day=1)

        fy_options = self._get_payroll_fiscal_year_options()

        return {
            'selected_month': selected_month,
            'selected_fy_start': fy_start,
            'selected_fy_end': fy_end,
            'selected_fy_value': selected_fy_value,
            'selected_fy_code': selected_fy_code,
            'selected_fy_label': selected_fy_code,
            'selected_fy_range_label': f'{fy_start.day} {fy_start.strftime("%B %Y")} to {fy_end.day} {fy_end.strftime("%B %Y")}',
            'fiscal_year_options': fy_options,
            'fiscal_year_month_options': month_options,
        }

    # ---------------------------------------------------------
    # Secure Chatbot iframe routes
    # ---------------------------------------------------------
    def _get_chatbot_secret(self):
        return (config.get('chatbot_shared_secret') or '').strip()

    def _sign_chatbot_payload(self, payload):
        secret = self._get_chatbot_secret()
        if not secret:
            return ''
        return hmac.new(secret.encode('utf-8'), payload.encode('utf-8'), hashlib.sha256).hexdigest()

    @http.route('/my/hr/chatbot/employee/frame', type='http', auth='user', website=True)
    def my_hr_employee_chatbot_frame(self, **kwargs):
        hr_maintenance_response = self._redirect_if_hr_portal_maintenance()
        if hr_maintenance_response is not False:
            return hr_maintenance_response

        employee = request.env['hr.employee'].sudo().search([
            ('user_id', '=', request.env.user.id)
        ], limit=1)

        if not employee or not employee.employee_code:
            return request.make_response(
                "Employee chatbot access denied: employee record not found.",
                headers=[('Content-Type', 'text/plain')]
            )

        ts = str(int(pytime.time()))
        employee_code = employee.employee_code.strip()
        payload = f"employee|{employee_code}|{ts}"
        sig = self._sign_chatbot_payload(payload)

        if not sig:
            return request.make_response(
                "Employee chatbot security is not configured.",
                headers=[('Content-Type', 'text/plain')]
            )

        query = urlencode({
            'employee_code': employee_code,
            'ts': ts,
            'sig': sig,
        })
        return request.redirect('/chatbot/employee/?' + query)

    @http.route('/my/hr/chatbot/admin/frame', type='http', auth='user', website=True)
    def my_hr_admin_chatbot_frame(self, **kwargs):
        if not self._is_hr_manager():
            return request.make_response(
                "Admin chatbot access denied.",
                headers=[('Content-Type', 'text/plain')]
            )

        ts = str(int(pytime.time()))
        payload = f"admin|{ts}"
        sig = self._sign_chatbot_payload(payload)

        if not sig:
            return request.make_response(
                "Admin chatbot security is not configured.",
                headers=[('Content-Type', 'text/plain')]
            )

        query = urlencode({
            'ts': ts,
            'sig': sig,
        })
        return request.redirect('/chatbot/admin/?' + query)


    # ---------------------------------------------------------
    # Request helpers
    # ---------------------------------------------------------


    def _get_request_type_label(self, request_type):
        for option_value, option_label in self.REQUEST_TYPE_OPTIONS:
            if option_value == request_type:
                return option_label
        return ''

    def _parse_request_date(self, date_string):
        try:
            return datetime.strptime(date_string or '', '%Y-%m-%d').date()
        except ValueError:
            return False

    def _send_employee_request_email(self, employee, request_type, reason, date_from, date_to):
        if 'mail.mail' not in request.env:
            return False, 'Outgoing email is not available in this database yet.'

        request_type_label = self._get_request_type_label(request_type)
        employee_name = employee.name or 'Employee'
        employee_code = employee.employee_code or '-'
        department_name = employee.department_id.name or '-'
        designation_name = employee.job_id.name or '-'
        requester_email = request.env.user.email or employee.work_email or '-'

        mail_subject = f"HR Portal Request - {request_type_label} - {employee_name}"

        mail_body = f"""
            <div style="font-family: Arial, sans-serif; font-size: 14px; color: #222;">
                <h3 style="margin-bottom: 12px;">New HR Portal Request</h3>
                <table style="border-collapse: collapse; width: 100%; max-width: 700px;">
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;"><strong>Employee Name</strong></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{employee_name}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;"><strong>Employee Code</strong></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{employee_code}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;"><strong>Department</strong></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{department_name}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;"><strong>Designation</strong></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{designation_name}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;"><strong>Requester Email</strong></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{requester_email}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;"><strong>Request Type</strong></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{request_type_label}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;"><strong>From Date</strong></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{date_from}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;"><strong>To Date</strong></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{date_to}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd; vertical-align: top;"><strong>Reason</strong></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{reason}</td>
                    </tr>
                </table>
                <div style="margin-top: 20px;">
                    <a href="https://odoo.blimp.pk/my/hr/admin/leaves#hr-requests"
                       style="background-color: #006BB6; color: #ffffff; padding: 12px 18px; text-decoration: none; border-radius: 6px; font-weight: bold; display: inline-block;">
                        Respond to this request
                    </a>
                </div>
            </div>
        """

        mail_values = {
            'subject': mail_subject,
            'body_html': mail_body,
            'email_to': self.REQUEST_EMAIL_TO,
            'email_from': request.env.user.email or 'noreply@example.com',
            'reply_to': request.env.user.email or '',
        }

        mail = request.env['mail.mail'].sudo().create(mail_values)
        mail.send()

        return True, 'Your request has been sent successfully.'
    def _send_employee_request_status_email(self, portal_request):
        if 'mail.mail' not in request.env:
            return False

        employee = portal_request.employee_id
        employee_email = (
            employee.user_id.email
            or employee.work_email
            or request.env.user.email
            or ''
        )

        if not employee_email:
            return False

        status_label = 'Approved' if portal_request.state == 'approved' else 'Rejected'

        mail_subject = f"HR Request {status_label} - {portal_request.request_type_label}"

        mail_body = f"""
            <div style="font-family: Arial, sans-serif; font-size: 14px; color: #222;">
                <h3>HR Request {status_label}</h3>
                <p>Dear {employee.name or 'Employee'},</p>
                <p>Your HR request has been <strong>{status_label}</strong>.</p>

                <table style="border-collapse: collapse; width: 100%; max-width: 700px;">
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;"><strong>Request Type</strong></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{portal_request.request_type_label or '-'}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;"><strong>From Date</strong></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{portal_request.date_from or '-'}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;"><strong>To Date</strong></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{portal_request.date_to or '-'}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;"><strong>Admin Remarks</strong></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{portal_request.admin_remarks or '-'}</td>
                    </tr>
                </table>
            </div>
        """

        request.env['mail.mail'].sudo().create({
            'subject': mail_subject,
            'body_html': mail_body,
            'email_to': employee_email,
            'email_from': request.env.user.email or 'noreply@example.com',
        }).send()

        return True

    # ---------------------------------------------------------

    # ---------------------------------------------------------
    # HR Portal Maintenance Mode
    # ---------------------------------------------------------
    def _is_hr_portal_maintenance_enabled(self):
        return request.env['ir.config_parameter'].sudo().get_param(
            'hr_employee_portal.maintenance_mode', '0'
        ) == '1'

    def _redirect_if_hr_portal_maintenance(self):
        if not self._is_hr_portal_maintenance_enabled():
            return False

        current_path = request.httprequest.path or ''

        # Admin HR portal must remain available during maintenance.
        if current_path.startswith('/my/hr/admin'):
            return False

        # If an HR manager/admin opens an employee HR URL, send them back to admin.
        if self._is_hr_manager():
            return request.redirect('/my/hr/admin')

        # Employees are blocked from all employee HR portal pages.
        return request.render('hr_employee_portal.hr_employee_maintenance_page', {
            'page_name': 'hr_maintenance',
            'no_breadcrumbs': True,
        })


    def _get_valid_work_modes(self):
        return ['office', 'wfh', 'field', 'leave']

    def _build_redirect_url(self, base_path, params=None):
        params = params or {}
        clean_params = {}
        for key, value in params.items():
            if value in (False, None, ''):
                continue
            clean_params[key] = value

        if not clean_params:
            return base_path
        return '%s?%s' % (base_path, url_encode(clean_params))

    def _parse_portal_date(self, date_string):
        try:
            return datetime.strptime((date_string or '').strip(), '%Y-%m-%d').date()
        except ValueError:
            return False
    def _parse_portal_float(self, value):
        try:
            clean_value = str(value or '').replace(',', '').strip()
            if not clean_value:
                return 0.0
            return float(clean_value)
        except (TypeError, ValueError):
            return 0.0

    def _parse_portal_datetime_local(self, datetime_string):
        try:
            return datetime.strptime((datetime_string or '').strip(), '%Y-%m-%dT%H:%M:%S')
        except ValueError:
            try:
                return datetime.strptime((datetime_string or '').strip(), '%Y-%m-%dT%H:%M')
            except ValueError:
                return False



    def _expire_late_report_accesses(self):
        if 'hr.late.report.access' not in request.env:
            return

        expired_accesses = request.env['hr.late.report.access'].sudo().search([
            ('state', '=', 'active'),
            ('expires_at', '<=', fields.Datetime.now()),
        ])

        if expired_accesses:
            expired_accesses.write({'state': 'expired'})

    def _get_active_late_report_access(self, employee, report_date):
        self._expire_late_report_accesses()

        if 'hr.late.report.access' not in request.env or not employee or not report_date:
            return False

        today = fields.Date.context_today(request.env.user)
        yesterday = today - timedelta(days=1)
        if report_date >= yesterday:
            return False

        return request.env['hr.late.report.access'].sudo().search([
            ('employee_id', '=', employee.id),
            ('report_date', '=', report_date),
            ('state', '=', 'active'),
            ('expires_at', '>', fields.Datetime.now()),
        ], limit=1)

    def _get_active_late_report_accesses(self, employee):
        self._expire_late_report_accesses()

        if 'hr.late.report.access' not in request.env or not employee:
            return request.env['hr.late.report.access'].sudo().browse([])

        today = fields.Date.context_today(request.env.user)
        yesterday = today - timedelta(days=1)

        return request.env['hr.late.report.access'].sudo().search([
            ('employee_id', '=', employee.id),
            ('report_date', '<', yesterday),
            ('state', '=', 'active'),
            ('expires_at', '>', fields.Datetime.now()),
        ], order='report_date desc, expires_at asc')

    def _get_task_report_search_domain(self, employee_id=None, date_from=None, date_to=None):
        domain = []


        if employee_id:
            domain.append(('employee_id', '=', employee_id))

        if date_from:
            domain.append(('report_date', '>=', date_from))

        if date_to:
            domain.append(('report_date', '<=', date_to))

        return domain

    def _get_admin_performance_data(self, employee_id=None, date_from=None, date_to=None):
        performance_env = request.env['hr.daily.performance.plan'].sudo()
        employee_env = request.env['hr.employee'].sudo()

        domain = []

        if employee_id:
            domain.append(('employee_id', '=', employee_id))
        if date_from:
            domain.append(('plan_date', '>=', date_from))
        if date_to:
            domain.append(('plan_date', '<=', date_to))

        records = performance_env.search(domain, order='plan_date desc, id desc')
        employees = employee_env.search([], order='name asc')

        return {
            'employees': employees,
            'records': records,
        }

    def _get_valid_performance_priorities(self):
        return ['low', 'medium', 'high']

    def _get_valid_performance_statuses(self):
        return ['pending', 'underprocess', 'completed', 'hold']

    def _get_employee_performance_records(self, employee, date_from=None, date_to=None):
        if not employee:
            return request.env['hr.daily.performance.plan'].sudo()

        domain = [('employee_id', '=', employee.id)]

        if date_from:
            domain.append(('plan_date', '>=', date_from))
        if date_to:
            domain.append(('plan_date', '<=', date_to))

        return request.env['hr.daily.performance.plan'].sudo().search(
            domain,
            order='plan_date desc, id desc'
        )

    def _get_performance_search_domain(self, employee_id=None, date_from=None, date_to=None):
        domain = []

        if employee_id:
            domain.append(('employee_id', '=', employee_id))

        if date_from:
            domain.append(('plan_date', '>=', date_from))

        if date_to:
            domain.append(('plan_date', '<=', date_to))

        return domain

    def _get_all_performance_employees(self):
        return request.env['hr.employee'].sudo().search([], order='name asc')

        # ---------------------------------------------------------
    # Payroll helpers
    # ---------------------------------------------------------
    def _get_employee_payroll_period_label(self, selected_month):
        return selected_month.strftime('%B %Y')

    def _get_payroll_month_overview(self, selected_month):
        total_days_in_month = calendar.monthrange(selected_month.year, selected_month.month)[1]
        today = fields.Date.context_today(request.env.user)

        if selected_month.year == today.year and selected_month.month == today.month:
            elapsed_days = today.day
        else:
            elapsed_days = total_days_in_month

        payment_date = (selected_month + timedelta(days=32)).replace(day=5)
        fiscal_context = self._get_fiscal_year_context(selected_month)

        return {
            'month_label': selected_month.strftime('%b-%y'),
            'full_month_label': selected_month.strftime('%B %Y'),
            'payment_date': payment_date.strftime('%d-%b-%y'),
            'days_in_month_elapsed': elapsed_days,
            'days_in_month_total': total_days_in_month,
            'fiscal_year_label': fiscal_context['fiscal_year_label'],
            'fiscal_year_code': fiscal_context['fiscal_year_code'],
            'fiscal_year_range_label': fiscal_context['fiscal_year_range_label'],
        }

    def _get_payroll_register_line(self, employee, selected_month):
        if not employee or not selected_month:
            return request.env['hr.payroll.register.line']

        month_start = selected_month.replace(day=1)
        return request.env['hr.payroll.register.line'].sudo().search([
            ('employee_id', '=', employee.id),
            ('month_date', '=', month_start),
        ], limit=1)

    def _get_admin_payroll_register_lines(self, selected_month, employee_id=None):
        if not selected_month:
            return request.env['hr.payroll.register.line'].sudo()

        month_start = selected_month.replace(day=1)
        domain = [('month_date', '=', month_start)]

        if employee_id:
            domain.append(('employee_id', '=', employee_id))

        return request.env['hr.payroll.register.line'].sudo().search(
            domain,
            order='employee_id asc'
        )

    def _get_employee_payroll_data(self, employee, selected_month):
        payroll_line = self._get_payroll_register_line(employee, selected_month)
        payroll_overview = self._get_payroll_month_overview(selected_month)

        if not payroll_line:
            return {
                'period_label': self._get_employee_payroll_period_label(selected_month),
                'overview': payroll_overview,
                'fiscal_year_code': payroll_overview['fiscal_year_code'],
                'fiscal_year_range_label': payroll_overview['fiscal_year_range_label'],
                'has_payroll': False,
                'record': False,
                'earnings': {
                    'basic': 0.0,
                    'actual_basic_working': 0.0,
                    'medical_allowance': 0.0,
                    'advertised_salary': 0.0,
                    'project_salary': 0.0,
                    'bonus': 0.0,
                    'allowance': 0.0,
                    'total_earnings': 0.0,
                },
                'deductions': {
                    'income_tax': 0.0,
                    'other_deductions': 0.0,
                    'loans_from_cash': 0.0,
                    'total_deductions': 0.0,
                },
                'summary': {
                    'total': 0.0,
                    'total_round': 0.0,
                    'net_salary': 0.0,
                },
                'notes': {
                    'designation': '',
                    'payment_method': '',
                    'allowance_detail': '',
                    'period_label': '',
                    'paid_at': '',
                    'comments': '',
                    'payment_date': '',
                    'month_label': selected_month.strftime('%B %Y'),
                },
                'tax_slabs': self.ADMIN_PAYROLL_TAX_SLABS,
            }

        total_earnings = self._to_amount(
            payroll_line.basic_actual
            + payroll_line.medical_allowance
            + payroll_line.advertised_salary
            + payroll_line.project_salary
            + payroll_line.bonus
            + payroll_line.allowance
        )

        total_deductions = self._to_amount(
            payroll_line.income_tax_deduction
            + payroll_line.other_deductions
            + payroll_line.loans_from_cash
        )

        return {
            'period_label': payroll_line.month_label or self._get_employee_payroll_period_label(selected_month),
            'overview': payroll_overview,
            'fiscal_year_code': payroll_overview['fiscal_year_code'],
            'fiscal_year_range_label': payroll_overview['fiscal_year_range_label'],
            'has_payroll': True,
            'record': payroll_line,
            'earnings': {
                'basic': self._to_amount(payroll_line.basic_salary),
                'actual_basic_working': self._to_amount(payroll_line.basic_actual),
                'medical_allowance': self._to_amount(payroll_line.medical_allowance),
                'advertised_salary': self._to_amount(payroll_line.advertised_salary),
                'project_salary': self._to_amount(payroll_line.project_salary),
                'bonus': self._to_amount(payroll_line.bonus),
                'allowance': self._to_amount(payroll_line.allowance),
                'total_earnings': total_earnings,
            },
            'deductions': {
                'income_tax': self._to_amount(payroll_line.income_tax_deduction),
                'other_deductions': self._to_amount(payroll_line.other_deductions),
                'loans_from_cash': self._to_amount(payroll_line.loans_from_cash),
                'total_deductions': total_deductions,
            },
            'summary': {
                'total': self._to_amount(payroll_line.total),
                'total_round': self._to_amount(payroll_line.total_round),
                'net_salary': self._to_amount(payroll_line.total_salary),
            },
            'notes': {
                'designation': payroll_line.designation or '',
                'payment_method': payroll_line.payment_method or '',
                'allowance_detail': payroll_line.allowance_detail or '',
                'period_label': payroll_line.period_label or '',
                'paid_at': payroll_line.paid_at or '',
                'comments': payroll_line.comments or '',
                'payment_date': payroll_line.payment_date.strftime('%d-%b-%Y') if payroll_line.payment_date else '',
                'month_label': payroll_line.month_label or '',
            },
            'tax_slabs': self.ADMIN_PAYROLL_TAX_SLABS,
        }

    def _sum_row_values(self, rows, key):
        return self._to_amount(sum(self._to_amount(row.get(key, 0.0)) for row in rows))
    def _export_workbook_response(self, workbook, file_name):
        for sheet in workbook.worksheets:
            for column_cells in sheet.columns:
                max_length = 12
                column_letter = column_cells[0].column_letter

                for cell in column_cells:
                    value = str(cell.value or '')
                    max_length = max(max_length, min(len(value) + 2, 45))

                sheet.column_dimensions[column_letter].width = max_length

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        return request.make_response(
            output.getvalue(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', 'attachment; filename="%s"' % file_name),
            ]
        )

    def _count_weekend_days(self, month_start, end_day):
        weekend_days = 0
        for day_number in range(1, end_day + 1):
            target_date = month_start.replace(day=day_number)
            if target_date.weekday() >= 5:
                weekend_days += 1
        return weekend_days

    def _get_admin_payroll_overview(self, selected_month):
        total_days_in_month = calendar.monthrange(selected_month.year, selected_month.month)[1]
        today = fields.Date.context_today(request.env.user)

        if selected_month.year == today.year and selected_month.month == today.month:
            elapsed_days = today.day
        else:
            elapsed_days = total_days_in_month

        holiday_weekend_days = self._count_weekend_days(selected_month, elapsed_days)
        work_days = max(elapsed_days - holiday_weekend_days, 0)

        payment_date = (selected_month + timedelta(days=32)).replace(day=5)
        fiscal_context = self._get_fiscal_year_context(selected_month)

        return {
            'month_label': selected_month.strftime('%b-%y'),
            'full_month_label': selected_month.strftime('%B %Y'),
            'payment_date': payment_date.strftime('%d-%b-%y'),
            'fiscal_year_label': fiscal_context['fiscal_year_label'],
            'fiscal_year_code': fiscal_context['fiscal_year_code'],
            'fiscal_year_range_label': fiscal_context['fiscal_year_range_label'],
            'work_days': work_days,
            'holiday_weekend_days': holiday_weekend_days,
            'days_in_month_elapsed': elapsed_days,
            'days_in_month_total': total_days_in_month,
        }

    def _get_admin_payroll_data(self, selected_month, employee_id=None):
        payroll_lines = self._get_admin_payroll_register_lines(selected_month, employee_id=employee_id)

        rows = []
        for line in payroll_lines:
                project_salary_value = self._to_amount(getattr(line, 'project_salary', 0.0))
                allowance_value = self._to_amount(getattr(line, 'allowance', 0.0))
                project_salary_allowance = self._to_amount(project_salary_value + allowance_value)

                project_detail = (
                    getattr(line, 'project_detail', False)
                    or getattr(line, 'allowance_detail', False)
                    or '-'
        )
                bonus_detail = getattr(line, 'bonus_detail', False) or '-'
                deduction_detail = (
                    getattr(line, 'deduction_detail', False)
                    or getattr(line, 'comments', False)
                    or '-'
        )

                rows.append({
                    'record_id': line.id,
                    'employee_code': line.employee_code or '-',
                    'employee_name': line.employee_id.name or line.employee_name_text or '-',
                    'bank_name': line.payment_method or '-',
                    'designation': line.designation or '-',
                    'payment_method': line.payment_method or '-',

                    'basic_salary': self._to_amount(line.basic_salary),
                    'basic_actual': self._to_amount(line.basic_actual),
                    'medical_allowance': self._to_amount(line.medical_allowance),
                    'advertised_salary': self._to_amount(line.advertised_salary),

                    'project_salary': project_salary_allowance,
                    'project_value': project_detail,

                    'bonus': self._to_amount(line.bonus),
                    'for_value': bonus_detail,

                    'overtime': self._to_amount(getattr(line, 'overtime', 0.0)),
                    'ot_detail': getattr(line, 'overtime_detail', False) or '-',

                    'taxable_income': self._to_amount(line.taxable_income),
                    'yearly_income': self._to_amount(line.yearly_income),
                    'income_tax_deduction': self._to_amount(line.income_tax_deduction),
                    'other_deductions': self._to_amount(line.other_deductions),
                    'deduction_for': deduction_detail,

                    'total': self._to_amount(line.total),
                    'total_round': self._to_amount(line.total_round),
                    'total_allowance': allowance_value,
                    'hr_cost_including_bonus': self._to_amount(line.total_salary),
                    'hr_cost_rounded': self._to_amount(line.total_round),

                    'payment_date': line.payment_date.strftime('%d-%b-%Y') if line.payment_date else '-',
                    'payment_date_value': line.payment_date.strftime('%Y-%m-%d') if line.payment_date else '',
                    'month_label': line.month_label or '-',

                    'project_detail_raw': (
                        getattr(line, 'project_detail', False)
                        or getattr(line, 'allowance_detail', False)
                        or ''
            ),
                    'bonus_detail_raw': getattr(line, 'bonus_detail', False) or '',
                    'deduction_detail_raw': getattr(line, 'deduction_detail', False) or '',
                    'comments_raw': line.comments or '',
        })

        totals = {
                'basic_salary': self._sum_row_values(rows, 'basic_salary'),
                'basic_actual': self._sum_row_values(rows, 'basic_actual'),
                'medical_allowance': self._sum_row_values(rows, 'medical_allowance'),
                'advertised_salary': self._sum_row_values(rows, 'advertised_salary'),
                'project_salary': self._sum_row_values(rows, 'project_salary'),
                'project_value': 0.0,
                'bonus': self._sum_row_values(rows, 'bonus'),
                'for_value': 0.0,
                'overtime': self._sum_row_values(rows, 'overtime'),
                'ot_detail': 0.0,
                'taxable_income': self._sum_row_values(rows, 'taxable_income'),
                'yearly_income': self._sum_row_values(rows, 'yearly_income'),
                'income_tax_deduction': self._sum_row_values(rows, 'income_tax_deduction'),
                'other_deductions': self._sum_row_values(rows, 'other_deductions'),
                'deduction_for': 0.0,
                'total': self._sum_row_values(rows, 'total'),
                'total_round': self._sum_row_values(rows, 'total_round'),
                'total_allowance': self._sum_row_values(rows, 'total_allowance'),
                'hr_cost_including_bonus': self._sum_row_values(rows, 'hr_cost_including_bonus'),
                'hr_cost_rounded': self._sum_row_values(rows, 'hr_cost_rounded'),
    }

        calculator_input = rows[0]['total_round'] if rows else 0.0
        calculator_medical = rows[0]['medical_allowance'] if rows else 0.0
        calculator_taxable_yearly = self._to_amount(calculator_input * 12.0)
        calculator_monthly_tax = rows[0]['income_tax_deduction'] if rows else 0.0
        calculator_yearly_tax = self._to_amount(calculator_monthly_tax * 12.0)

        overview = self._get_admin_payroll_overview(selected_month)

        return {
            'overview': overview,
            'fiscal_year_code': overview['fiscal_year_code'],
            'fiscal_year_range_label': overview['fiscal_year_range_label'],
            'rows': rows,
            'totals': totals,
            'tax_slabs': self.ADMIN_PAYROLL_TAX_SLABS,
            'calculator': {
                'put_net_salary': calculator_input,
                'medical_allowance': calculator_medical,
                'taxable_income_yearly': calculator_taxable_yearly,
                'monthly_tax': calculator_monthly_tax,
                'yearly_tax': calculator_yearly_tax,
            },
        }

    # ---------------------------------------------------------
    # Get leave code for a specific date
    # ---------------------------------------------------------
    def _get_leave_code_for_date(self, employee, target_date):
        leave_domain = [
            ('employee_id', '=', employee.id),
            ('state', 'in', ['validate', 'validate1']),
            ('request_date_from', '<=', target_date),
            ('request_date_to', '>=', target_date),
        ]
        leave = request.env['hr.leave'].sudo().search(leave_domain, limit=1)

        if not leave:
            target_start = datetime.combine(target_date, time.min)
            target_end = datetime.combine(target_date, time.max)
            leave = request.env['hr.leave'].sudo().search([
                ('employee_id', '=', employee.id),
                ('state', 'in', ['validate', 'validate1']),
                ('date_from', '<=', target_end),
                ('date_to', '>=', target_start),
            ], limit=1)

        if not leave:
            return False

        leave_name = (leave.holiday_status_id.name or '').strip().lower()

        if 'sick' in leave_name:
            return 'S'
        if 'casual' in leave_name:
            return 'C'
        if 'unpaid' in leave_name:
            return 'U'

        return 'U'

    # ---------------------------------------------------------
    # Check holiday/weekend/public holiday
    # ---------------------------------------------------------
    def _is_holiday_for_date(self, employee, target_date):
        if target_date.weekday() >= 5:
            return True

        if 'resource.calendar.leaves' in request.env:
            day_start = datetime.combine(target_date, time.min)
            day_end = datetime.combine(target_date, time.max)

            holiday_domain = [
                ('date_from', '<=', day_end),
                ('date_to', '>=', day_start),
                '|',
                ('resource_id', '=', False),
                ('resource_id', '=', employee.resource_id.id),
            ]
            holiday_exists = request.env['resource.calendar.leaves'].sudo().search_count(holiday_domain)
            if holiday_exists:
                return True

        return False

    # ---------------------------------------------------------
    # Get daily work report record for a specific date
    # ---------------------------------------------------------
    def _get_daily_report_for_date(self, employee, target_date):
        if 'hr.daily.work.report' not in request.env:
            return False

        report = request.env['hr.daily.work.report'].sudo().search([
            ('employee_id', '=', employee.id),
            ('report_date', '=', target_date),
            ('state', 'in', ['submitted', 'approved']),
        ], limit=1)

        return report or False

    # ---------------------------------------------------------
    # Check if submitted late
    # ---------------------------------------------------------
    def _is_late_submission(self, report):
        if not report or not report.create_date or not report.report_date:
            return False

        timestamp_source = report.submitted_at or report.create_date
        local_dt = fields.Datetime.context_timestamp(report, timestamp_source)
        report_date = report.report_date

        if local_dt.date() > report_date:
            return True

        if local_dt.date() == report_date and local_dt.hour >= 18:
            return True

        return False
    def _get_attendance_register_map(self, employee, month_start):
        register_map = {}
        overtime_days = set()

        if 'hr.attendance.register.line' not in request.env or not employee:
            return register_map, overtime_days

        total_days = calendar.monthrange(month_start.year, month_start.month)[1]
        month_end = month_start.replace(day=total_days)

        register_lines = request.env['hr.attendance.register.line'].sudo().search([
            ('employee_id', '=', employee.id),
            ('attendance_date', '>=', month_start),
            ('attendance_date', '<=', month_end),
        ], order='attendance_date asc')

        for line in register_lines:
            register_map[line.attendance_date] = line.attendance_code
            if line.attendance_code == 'OT':
                overtime_days.add(line.attendance_date)

        return register_map, overtime_days

    # ---------------------------------------------------------
    # Build attendance matrix row for selected month
    # ---------------------------------------------------------
    def _build_attendance_matrix(self, employee, month_start):
        total_days = calendar.monthrange(month_start.year, month_start.month)[1]
        days = []
        summary = {
            'total': total_days,
            'P': 0,
            'R': 0,
            'H': 0,
            'S': 0,
            'C': 0,
            'U': 0,
            'D': 0,
            'OT': 0,
        }

        register_map, overtime_days = self._get_attendance_register_map(employee, month_start)

        for day_number in range(1, total_days + 1):
            target_date = month_start.replace(day=day_number)
            code = register_map.get(target_date, '-')
            title_parts = [target_date.strftime('%d %b %Y')]

            if code == 'P':
                title_parts.append('Present')
            elif code == 'R':
                title_parts.append('Remote / WFH')
            elif code == 'H':
                title_parts.append('Holiday')
            elif code == 'S':
                title_parts.append('Sick Leave')
            elif code == 'C':
                title_parts.append('Casual Leave')
            elif code == 'U':
                title_parts.append('Unpaid Leave')
            elif code == 'D':
                title_parts.append('Late Submission')
            elif code == 'OT':
                title_parts.append('Overtime')
            else:
                title_parts.append('No Record')

            if code in summary:
                summary[code] += 1

            days.append({
                'day_number': day_number,
                'weekday_short': target_date.strftime('%a'),
                'code': code,
                'title': ' | '.join(title_parts),
                'is_weekend': target_date.weekday() >= 5,
            })

        return {
            'fiscal_year': self._get_fiscal_year_label(month_start),
            'fiscal_year_code': self._get_fiscal_year_code(month_start),
            'month_label': month_start.strftime('%b %Y'),
            'employee_name': employee.name or '-',
            'days': days,
            'summary': summary,
        }
    # ---------------------------------------------------------
    # Leave helpers
    # ---------------------------------------------------------
    def _get_leave_date_range(self, leave):
        date_from = leave.request_date_from or False
        date_to = leave.request_date_to or False

        if not date_from and leave.date_from:
            date_from = leave.date_from.date()
        if not date_to and leave.date_to:
            date_to = leave.date_to.date()

        return date_from, date_to

    def _calculate_overlap_days(self, range_start, range_end, month_start, month_end):
        overlap_start = max(range_start, month_start)
        overlap_end = min(range_end, month_end)
        if overlap_start > overlap_end:
            return 0.0
        return float((overlap_end - overlap_start).days + 1)

    def _sum_leave_days_for_month(self, employee, month_start, leave_keyword):
        total_days = 0.0
        month_end = month_start.replace(
            day=calendar.monthrange(month_start.year, month_start.month)[1]
        )

        leave_records = request.env['hr.leave'].sudo().search([
            ('employee_id', '=', employee.id),
            ('state', 'in', ['validate', 'validate1']),
            ('request_date_from', '<=', month_end),
            ('request_date_to', '>=', month_start),
        ])

        for leave in leave_records:
            leave_name = (leave.holiday_status_id.name or '').strip().lower()
            if leave_keyword not in leave_name:
                continue

            leave_start, leave_end = self._get_leave_date_range(leave)
            if not leave_start or not leave_end:
                continue

            total_days += self._calculate_overlap_days(leave_start, leave_end, month_start, month_end)

        return round(total_days, 2)

    # ---------------------------------------------------------
    # WFH / Late data helpers
    # ---------------------------------------------------------
    def _count_wfh_for_month(self, employee, month_start):
        if 'hr.daily.work.report' not in request.env:
            return 0

        month_end = month_start.replace(
            day=calendar.monthrange(month_start.year, month_start.month)[1]
        )

        return request.env['hr.daily.work.report'].sudo().search_count([
            ('employee_id', '=', employee.id),
            ('report_date', '>=', month_start),
            ('report_date', '<=', month_end),
            ('work_mode', '=', 'wfh'),
            ('state', 'in', ['submitted', 'approved']),
        ])

    def _count_late_reports_for_month(self, employee, month_start):
        if 'hr.daily.work.report' not in request.env:
            return 0

        month_end = month_start.replace(
            day=calendar.monthrange(month_start.year, month_start.month)[1]
        )

        reports = request.env['hr.daily.work.report'].sudo().search([
            ('employee_id', '=', employee.id),
            ('report_date', '>=', month_start),
            ('report_date', '<=', month_end),
            ('state', 'in', ['submitted', 'approved']),
        ])

        return len(reports.filtered(lambda rec: self._is_late_submission(rec)))

    # ---------------------------------------------------------
    # FY attendance summary ledger
    # ---------------------------------------------------------
    def _build_attendance_summary_ledger(self, employee, selected_month):
        months = self._get_fiscal_year_month_starts(selected_month)
        fiscal_year_label = self._get_fiscal_year_label(selected_month)

        sick_opening = 0.0
        casual_opening = 0.0
        rows = []

        total_sick_availed = 0.0
        total_sick_earned = 0.0
        total_casual_availed = 0.0
        total_casual_earned = 0.0
        total_wfh = 0
        total_late_data = 0

        for month_start in months:
            month_is_active = month_start <= selected_month

            sick_earned = self.SICK_MONTHLY_EARN if month_is_active else 0.0
            casual_earned = self.CASUAL_MONTHLY_EARN if month_is_active else 0.0

            sick_availed = self._sum_leave_days_for_month(employee, month_start, 'sick') if month_is_active else 0.0
            casual_availed = self._sum_leave_days_for_month(employee, month_start, 'casual') if month_is_active else 0.0
            wfh_count = self._count_wfh_for_month(employee, month_start) if month_is_active else 0
            late_data_count = self._count_late_reports_for_month(employee, month_start) if month_is_active else 0

            sick_closing = round(sick_opening + sick_earned - sick_availed, 2)
            casual_closing = round(casual_opening + casual_earned - casual_availed, 2)

            rows.append({
                'month_label': month_start.strftime('%b %Y'),
                'month_value': month_start.strftime('%Y-%m'),
                'sick_opening': round(sick_opening, 2),
                'sick_availed': round(sick_availed, 2),
                'sick_earned': round(sick_earned, 2),
                'sick_closing': round(sick_closing, 2),
                'casual_opening': round(casual_opening, 2),
                'casual_availed': round(casual_availed, 2),
                'casual_earned': round(casual_earned, 2),
                'casual_closing': round(casual_closing, 2),
                'wfh_count': wfh_count,
                'late_data_count': late_data_count,
                'is_selected_month': month_start == selected_month,
            })

            total_sick_availed += sick_availed
            total_sick_earned += sick_earned
            total_casual_availed += casual_availed
            total_casual_earned += casual_earned
            total_wfh += wfh_count
            total_late_data += late_data_count

            sick_opening = sick_closing
            casual_opening = casual_closing

        balance_row = {
            'sick_opening': round(rows[-1]['sick_opening'] if rows else 0.0, 2),
            'sick_availed': round(total_sick_availed, 2),
            'sick_earned': round(total_sick_earned, 2),
            'sick_closing': round(rows[-1]['sick_closing'] if rows else 0.0, 2),
            'casual_opening': round(rows[-1]['casual_opening'] if rows else 0.0, 2),
            'casual_availed': round(total_casual_availed, 2),
            'casual_earned': round(total_casual_earned, 2),
            'casual_closing': round(rows[-1]['casual_closing'] if rows else 0.0, 2),
            'wfh_count': total_wfh,
            'late_data_count': total_late_data,
        }

        return {
            'fiscal_year_label': fiscal_year_label,
            'fiscal_year_code': self._get_fiscal_year_code(selected_month),
            'fiscal_year_range_label': self._get_fiscal_year_range_label(selected_month),
            'policy_sick_leaves': self.SICK_POLICY_DAYS,
            'policy_casual_leaves': self.CASUAL_POLICY_DAYS,
            'rows': rows,
            'balance_row': balance_row,
        }
    def _get_present_today_from_register(self):
        today = fields.Date.context_today(request.env.user)

        register_lines = request.env['hr.attendance.register.line'].sudo().search([
            ('attendance_date', '=', today),
            ('attendance_code', 'in', ['P', 'R']),
    ])

        employee_ids = register_lines.mapped('employee_id').ids
        return len(set(employee_ids))

    # ---------------------------------------------------------
    # Attendance logs helper
    # ---------------------------------------------------------
    def _get_attendance_logs(self, employee):
        attendance_display = []
        if not employee:
            return attendance_display

        attendances = request.env['hr.attendance'].sudo().search(
            [('employee_id', '=', employee.id)],
            order='check_in desc'
        )

        for att in attendances:
            attendance_display.append({
                'check_in': att.check_in,
                'check_out': att.check_out,
                'worked_hours': self._format_hours(att.worked_hours),
            })

        return attendance_display
    @http.route('/my/hr/performance', type='http', auth='user', website=True)
    def my_hr_performance(self, **kwargs):
        hr_maintenance_response = self._redirect_if_hr_portal_maintenance()
        if hr_maintenance_response is not False:
            return hr_maintenance_response

        if self._is_hr_manager():
            return request.redirect('/my/hr/admin/performance')

        employee = self._get_employee()
        redirect_response = self._redirect_if_no_employee(employee)
        if redirect_response:
            return redirect_response

        date_from = self._parse_portal_date(kwargs.get('date_from'))
        date_to = self._parse_portal_date(kwargs.get('date_to'))

        try:
            page = int(kwargs.get('page') or 1)
        except (TypeError, ValueError):
            page = 1

        page = max(page, 1)
        per_page = 8
        offset = (page - 1) * per_page

        performance_env = request.env['hr.daily.performance.plan'].sudo()

        domain = [('employee_id', '=', employee.id)]
        if date_from:
            domain.append(('plan_date', '>=', date_from))
        if date_to:
            domain.append(('plan_date', '<=', date_to))

        performance_total = performance_env.search_count(domain)

        performance_records = performance_env.search(
            domain,
            order='plan_date desc, id desc',
            limit=per_page,
            offset=offset,
        )

        performance_page_count = max((performance_total + per_page - 1) // per_page, 1)
        performance_start = offset + 1 if performance_total else 0
        performance_end = min(offset + per_page, performance_total)

        base_params = {
            'date_from': date_from.strftime('%Y-%m-%d') if date_from else '',
            'date_to': date_to.strftime('%Y-%m-%d') if date_to else '',
        }

        previous_page_url = False
        next_page_url = False

        if page > 1:
            previous_params = dict(base_params)
            previous_params['page'] = page - 1
            previous_page_url = self._build_redirect_url('/my/hr/performance', previous_params)

        if page < performance_page_count:
            next_params = dict(base_params)
            next_params['page'] = page + 1
            next_page_url = self._build_redirect_url('/my/hr/performance', next_params)

        values = self._prepare_portal_values(employee, {
            'performance_records': performance_records,
            'date_from': date_from.strftime('%Y-%m-%d') if date_from else '',
            'date_to': date_to.strftime('%Y-%m-%d') if date_to else '',
            'performance_status': (kwargs.get('performance_status') or '').strip(),
            'performance_message': (kwargs.get('performance_message') or '').strip(),
            'performance_total': performance_total,
            'performance_page': page,
            'performance_page_count': performance_page_count,
            'performance_start': performance_start,
            'performance_end': performance_end,
            'previous_page_url': previous_page_url,
            'next_page_url': next_page_url,
        })
        return request.render('hr_employee_portal.hr_employee_performance_page', values)
    @http.route('/my/hr/performance/submit', type='http', auth='user', methods=['POST'], website=True)
    def my_hr_performance_submit(self, **post):
        hr_maintenance_response = self._redirect_if_hr_portal_maintenance()
        if hr_maintenance_response is not False:
            return hr_maintenance_response

        if self._is_hr_manager():
            return request.redirect('/my/hr/admin/performance')

        employee = self._get_employee()
        if not employee:
            redirect_url = self._build_redirect_url('/my/hr', {
                'performance_status': 'error',
                'performance_message': 'Employee record not found.',
            })
            return request.redirect(redirect_url)

        if 'hr.daily.performance.plan' not in request.env:
            redirect_url = self._build_redirect_url('/my/hr/performance', {
                'performance_status': 'error',
                'performance_message': 'Performance module is not available.',
            })
            return request.redirect(redirect_url)

        plan_date = self._parse_portal_date(post.get('plan_date'))
        task_id = (post.get('task_id') or '').strip()
        project_name = (post.get('project_name') or '').strip()
        task_description = (post.get('task_description') or '').strip()
        priority_level = (post.get('priority_level') or '').strip()
        status = (post.get('status') or '').strip()

        if not plan_date:
            redirect_url = self._build_redirect_url('/my/hr/performance', {
                'performance_status': 'error',
                'performance_message': 'Please enter a valid date.',
            })
            return request.redirect(redirect_url)

        if not task_id or not project_name or not task_description:
            redirect_url = self._build_redirect_url('/my/hr/performance', {
                'performance_status': 'error',
                'performance_message': 'Task ID, Project, and Task Description are required.',
            })
            return request.redirect(redirect_url)

        if priority_level not in self._get_valid_performance_priorities():
            redirect_url = self._build_redirect_url('/my/hr/performance', {
                'performance_status': 'error',
                'performance_message': 'Please select a valid priority level.',
            })
            return request.redirect(redirect_url)

        if status not in self._get_valid_performance_statuses():
            redirect_url = self._build_redirect_url('/my/hr/performance', {
                'performance_status': 'error',
                'performance_message': 'Please select a valid status.',
            })
            return request.redirect(redirect_url)

        try:
            request.env['hr.daily.performance.plan'].sudo().create({
                'employee_id': employee.id,
                'plan_date': plan_date,
                'task_id': task_id,
                'project_name': project_name,
                'task_description': task_description,
                'priority_level': priority_level,
                'status': status,
            })
        except ValidationError as error:
            redirect_url = self._build_redirect_url('/my/hr/performance', {
                'performance_status': 'error',
                'performance_message': str(error),
            })
            return request.redirect(redirect_url)
        except Exception:
            redirect_url = self._build_redirect_url('/my/hr/performance', {
                'performance_status': 'error',
                'performance_message': 'Performance record could not be created.',
            })
            return request.redirect(redirect_url)

        redirect_url = self._build_redirect_url('/my/hr/performance', {
            'performance_status': 'success',
            'performance_message': 'Performance record created successfully.',
        })
        return request.redirect(redirect_url)

    @http.route('/my/hr/performance/update', type='http', auth='user', methods=['POST'], website=True)
    def my_hr_performance_update(self, **post):
        hr_maintenance_response = self._redirect_if_hr_portal_maintenance()
        if hr_maintenance_response is not False:
            return hr_maintenance_response

        if self._is_hr_manager():
            return request.redirect('/my/hr/admin/performance')

        employee = self._get_employee()
        if not employee:
            return request.redirect('/my/hr')

        record_id_raw = (post.get('record_id') or '').strip()
        try:
            record_id = int(record_id_raw)
        except (TypeError, ValueError):
            redirect_url = self._build_redirect_url('/my/hr/performance', {
                'performance_status': 'error',
                'performance_message': 'Invalid performance record.',
            })
            return request.redirect(redirect_url)

        record = request.env['hr.daily.performance.plan'].sudo().browse(record_id).exists()
        if not record or record.employee_id.id != employee.id:
            redirect_url = self._build_redirect_url('/my/hr/performance', {
                'performance_status': 'error',
                'performance_message': 'You can only edit your own performance records.',
            })
            return request.redirect(redirect_url)

        plan_date = self._parse_portal_date(post.get('plan_date'))
        task_id = (post.get('task_id') or '').strip()
        project_name = (post.get('project_name') or '').strip()
        task_description = (post.get('task_description') or '').strip()
        priority_level = (post.get('priority_level') or '').strip()
        status = (post.get('status') or '').strip()

        if not plan_date or not task_id or not project_name or not task_description:
            redirect_url = self._build_redirect_url('/my/hr/performance', {
                'performance_status': 'error',
                'performance_message': 'All editable fields are required.',
            })
            return request.redirect(redirect_url)

        if priority_level not in self._get_valid_performance_priorities():
            redirect_url = self._build_redirect_url('/my/hr/performance', {
                'performance_status': 'error',
                'performance_message': 'Invalid priority level selected.',
            })
            return request.redirect(redirect_url)

        if status not in self._get_valid_performance_statuses():
            redirect_url = self._build_redirect_url('/my/hr/performance', {
                'performance_status': 'error',
                'performance_message': 'Invalid status selected.',
            })
            return request.redirect(redirect_url)

        try:
            record.with_context(employee_portal_edit=True).write({
                'plan_date': plan_date,
                'task_id': task_id,
                'project_name': project_name,
                'task_description': task_description,
                'priority_level': priority_level,
                'status': status,
            })
        except ValidationError as error:
            redirect_url = self._build_redirect_url('/my/hr/performance', {
                'performance_status': 'error',
                'performance_message': str(error),
            })
            return request.redirect(redirect_url)
        except Exception:
            redirect_url = self._build_redirect_url('/my/hr/performance', {
                'performance_status': 'error',
                'performance_message': 'Performance record could not be updated.',
            })
            return request.redirect(redirect_url)

        redirect_url = self._build_redirect_url('/my/hr/performance', {
            'performance_status': 'success',
            'performance_message': 'Performance record updated successfully.',
        })
        return request.redirect(redirect_url)

    # =========================================================
    # EMPLOYEE: Dashboard Request Submit
    # =========================================================
    @http.route('/my/hr/request/submit', type='http', auth='user', website=True, methods=['POST'])
    def my_hr_request_submit(self, **post):
        hr_maintenance_response = self._redirect_if_hr_portal_maintenance()
        if hr_maintenance_response is not False:
            return hr_maintenance_response

        if self._is_hr_manager():
            return request.redirect('/my/hr/admin')

        employee = self._get_employee()
        maintenance_response = self._redirect_if_hr_portal_maintenance()
        if maintenance_response is not False:
            return maintenance_response
        if not employee:
            return request.redirect(self._build_redirect_url('/my/hr', {
                'request_status': 'error',
                'request_message': 'Employee record not found.',
            }))

        if 'hr.employee.portal.request' not in request.env:
            return request.redirect(self._build_redirect_url('/my/hr', {
                'request_status': 'error',
                'request_message': 'HR request model is not available. Please upgrade the module.',
            }))

        request_type = (post.get('request_type') or '').strip()
        reason = (post.get('reason') or '').strip()
        date_from_raw = (post.get('date_from') or '').strip()
        date_to_raw = (post.get('date_to') or '').strip()

        valid_request_types = [item[0] for item in self.REQUEST_TYPE_OPTIONS]

        if request_type not in valid_request_types:
            return request.redirect(self._build_redirect_url('/my/hr', {
                'request_status': 'error',
                'request_message': 'Please select a valid request type.',
            }))

        if not reason:
            return request.redirect(self._build_redirect_url('/my/hr', {
                'request_status': 'error',
                'request_message': 'Reason is required.',
            }))

        date_from = self._parse_request_date(date_from_raw)
        date_to = self._parse_request_date(date_to_raw)

        if not date_from or not date_to:
            return request.redirect(self._build_redirect_url('/my/hr', {
                'request_status': 'error',
                'request_message': 'Please enter valid from and to dates.',
            }))

        if date_to < date_from:
            return request.redirect(self._build_redirect_url('/my/hr', {
                'request_status': 'error',
                'request_message': 'To date cannot be earlier than from date.',
            }))

        try:
            portal_request = request.env['hr.employee.portal.request'].sudo().create({
                'employee_id': employee.id,
                'request_type': request_type,
                'date_from': date_from,
                'date_to': date_to,
                'reason': reason,
                'state': 'submitted',
                'submitted_at': fields.Datetime.now(),
                'late_submission_source': late_submission_source,
            })
        except Exception:
            return request.redirect(self._build_redirect_url('/my/hr', {
                'request_status': 'error',
                'request_message': 'Request could not be saved. Please contact HR.',
            }))

        try:
            self._send_employee_request_email(
                employee=employee,
                request_type=request_type,
                reason=reason,
                date_from=date_from,
                date_to=date_to,
            )
        except Exception:
            pass

        return request.redirect(self._build_redirect_url('/my/hr', {
            'request_status': 'success',
            'request_message': 'Request submitted successfully and sent to HR.',
        }))

    # =========================================================
    # EMPLOYEE: HR Dashboard
    # =========================================================
    @http.route('/my/hr', type='http', auth='user', website=True)
    def my_hr_dashboard(self, **kwargs):
        # Hard maintenance guard for employee dashboard route.
        if self._is_hr_portal_maintenance_enabled():
            if self._is_hr_manager():
                return request.redirect('/my/hr/admin')
            return request.render('hr_employee_portal.hr_employee_maintenance_page', {
                'page_name': 'hr_maintenance',
                'no_breadcrumbs': True,
            })

        if self._is_hr_manager():
            return request.redirect('/my/hr/admin')

        employee = self._get_employee()
        attendance_count = 0
        leave_count = 0

        if employee:
            attendance_count = request.env['hr.attendance'].sudo().search_count([
                ('employee_id', '=', employee.id)
            ])
            leave_count = request.env['hr.leave'].sudo().search_count([
                ('employee_id', '=', employee.id)
            ])

        values = self._prepare_portal_values(employee, {
            'attendance_count': attendance_count,
            'leave_count': leave_count,
            'document_count': 0,
            'request_status': kwargs.get('request_status', ''),
            'request_message': kwargs.get('request_message', ''),
        })
        return request.render('hr_employee_portal.hr_employee_portal_page', values)

    # ---------------------------------------------------------
    # EMPLOYEE: Profile Pages
    # ---------------------------------------------------------
    @http.route('/my/hr/profile', type='http', auth='user', website=True)
    def my_hr_profile(self, **kwargs):
        hr_maintenance_response = self._redirect_if_hr_portal_maintenance()
        if hr_maintenance_response is not False:
            return hr_maintenance_response

        if self._is_hr_manager():
            return request.redirect('/my/hr/admin')

        employee = self._get_employee()
        maintenance_response = self._redirect_if_hr_portal_maintenance()
        if maintenance_response is not False:
            return maintenance_response
        missing_employee = True if kwargs.get('missing_employee') else False
        values = self._prepare_portal_values(employee, {
            'missing_employee': missing_employee,
        })
        return request.render('hr_employee_portal.hr_employee_profile_page', values)

    @http.route('/my/hr/profile/employee-record', type='http', auth='user', website=True)
    def my_hr_profile_employee_record(self, **kwargs):
        hr_maintenance_response = self._redirect_if_hr_portal_maintenance()
        if hr_maintenance_response is not False:
            return hr_maintenance_response

        if self._is_hr_manager():
            return request.redirect('/my/hr/admin')

        employee = self._get_employee()
        maintenance_response = self._redirect_if_hr_portal_maintenance()
        if maintenance_response is not False:
            return maintenance_response
        redirect_response = self._redirect_if_no_employee(employee)
        if redirect_response:
            return redirect_response

        return request.render(
            'hr_employee_portal.hr_employee_profile_employee_record_page',
            self._prepare_portal_values(employee)
        )

    @http.route('/my/hr/profile/bank-account', type='http', auth='user', website=True)
    def my_hr_profile_bank_account(self, **kwargs):
        hr_maintenance_response = self._redirect_if_hr_portal_maintenance()
        if hr_maintenance_response is not False:
            return hr_maintenance_response

        if self._is_hr_manager():
            return request.redirect('/my/hr/admin')

        employee = self._get_employee()
        maintenance_response = self._redirect_if_hr_portal_maintenance()
        if maintenance_response is not False:
            return maintenance_response
        redirect_response = self._redirect_if_no_employee(employee)
        if redirect_response:
            return redirect_response

        return request.render(
            'hr_employee_portal.hr_employee_profile_bank_account_page',
            self._prepare_portal_values(employee)
        )

    @http.route('/my/hr/profile/personal-details', type='http', auth='user', website=True)
    def my_hr_profile_personal_details(self, **kwargs):
        hr_maintenance_response = self._redirect_if_hr_portal_maintenance()
        if hr_maintenance_response is not False:
            return hr_maintenance_response

        if self._is_hr_manager():
            return request.redirect('/my/hr/admin')

        employee = self._get_employee()
        maintenance_response = self._redirect_if_hr_portal_maintenance()
        if maintenance_response is not False:
            return maintenance_response
        redirect_response = self._redirect_if_no_employee(employee)
        if redirect_response:
            return redirect_response

        return request.render(
            'hr_employee_portal.hr_employee_profile_personal_details_page',
            self._prepare_portal_values(employee)
        )

    @http.route('/my/hr/profile/emergency-contact', type='http', auth='user', website=True)
    def my_hr_profile_emergency_contact(self, **kwargs):
        hr_maintenance_response = self._redirect_if_hr_portal_maintenance()
        if hr_maintenance_response is not False:
            return hr_maintenance_response

        if self._is_hr_manager():
            return request.redirect('/my/hr/admin')

        employee = self._get_employee()
        maintenance_response = self._redirect_if_hr_portal_maintenance()
        if maintenance_response is not False:
            return maintenance_response
        redirect_response = self._redirect_if_no_employee(employee)
        if redirect_response:
            return redirect_response

        return request.render(
            'hr_employee_portal.hr_employee_profile_emergency_contact_page',
            self._prepare_portal_values(employee)
        )

    @http.route('/my/hr/profile/employment-history', type='http', auth='user', website=True)
    def my_hr_profile_employment_history(self, **kwargs):
        hr_maintenance_response = self._redirect_if_hr_portal_maintenance()
        if hr_maintenance_response is not False:
            return hr_maintenance_response

        if self._is_hr_manager():
            return request.redirect('/my/hr/admin')

        employee = self._get_employee()
        maintenance_response = self._redirect_if_hr_portal_maintenance()
        if maintenance_response is not False:
            return maintenance_response
        redirect_response = self._redirect_if_no_employee(employee)
        if redirect_response:
            return redirect_response

        return request.render(
            'hr_employee_portal.hr_employee_profile_employment_history_page',
            self._prepare_portal_values(employee)
        )
    def _get_employee_attendance_fy_summary(self, employee, selected_month):
        summary = {
            'sick_opening': 0.0,
            'sick_availed': 0.0,
            'sick_earned': 0.0,
            'sick_closing': 0.0,
            'casual_opening': 0.0,
            'casual_availed': 0.0,
            'casual_earned': 0.0,
            'casual_closing': 0.0,
            'wfh_total': 0,
            'late_data_total': 0,
            'selected_month_code': selected_month.strftime('%Y-%m'),
            'selected_month_label': selected_month.strftime('%b %Y'),
        }

        if not employee:
            return summary

        payroll_period = self._get_selected_payroll_period(
            fy_value=None,
            month_value=selected_month.strftime('%Y-%m'),
        )
        fy_start = payroll_period['selected_fy_start']
        fy_end = payroll_period['selected_fy_end']

        register_env = request.env['hr.attendance.register.line'].sudo()

        selected_month_start = selected_month.replace(day=1)
        selected_month_end = selected_month_start.replace(
        day=calendar.monthrange(selected_month_start.year, selected_month_start.month)[1]
)

        register_lines = register_env.search([
        ('employee_id', '=', employee.id),
        ('attendance_date', '>=', selected_month_start),
        ('attendance_date', '<=', selected_month_end),
        ], order='attendance_date asc')

        leave_env = request.env['hr.leave'].sudo()
        sick_leaves = leave_env.search([
            ('employee_id', '=', employee.id),
            ('state', '=', 'validate'),
            ('request_date_from', '>=', fy_start),
            ('request_date_from', '<=', fy_end),
        ])

        sick_availed = 0.0
        casual_availed = 0.0
        for leave in sick_leaves:
            leave_name = (leave.holiday_status_id.name or '').strip().lower()
            days = leave.number_of_days or 0.0
            if 'sick' in leave_name:
                sick_availed += days
            elif 'casual' in leave_name:
                casual_availed += days

        sick_yearly_policy = 10.0
        casual_yearly_policy = 20.0

        sick_monthly_earned = round(sick_yearly_policy / 12.0, 2)
        casual_monthly_earned = round(casual_yearly_policy / 12.0, 2)

        months_elapsed = ((selected_month.year - fy_start.year) * 12) + (selected_month.month - fy_start.month) + 1
        months_elapsed = max(1, min(months_elapsed, 12))

        sick_earned_total = round(sick_monthly_earned * months_elapsed, 2)
        casual_earned_total = round(casual_monthly_earned * months_elapsed, 2)

        summary.update({
            'sick_opening': round(max(sick_earned_total - sick_monthly_earned, 0.0), 2),
            'sick_availed': round(sick_availed, 2),
            'sick_earned': sick_monthly_earned,
            'sick_closing': round(max(sick_earned_total - sick_availed, 0.0), 2),
            'casual_opening': round(max(casual_earned_total - casual_monthly_earned, 0.0), 2),
            'casual_availed': round(casual_availed, 2),
            'casual_earned': casual_monthly_earned,
            'casual_closing': round(max(casual_earned_total - casual_availed, 0.0), 2),
            'wfh_total': len(register_lines.filtered(lambda l: l.attendance_code == 'R')),
            'late_data_total': len(register_lines.filtered(lambda l: l.attendance_code == 'D')),
        })
        return summary

    # ---------------------------------------------------------
    # EMPLOYEE: Attendance Page
    # ---------------------------------------------------------
        # ---------------------------------------------------------
    # EMPLOYEE: Attendance Page
    # ---------------------------------------------------------
    @http.route('/my/hr/attendance', type='http', auth='user', website=True)
    def my_hr_attendance(self, **kwargs):
        hr_maintenance_response = self._redirect_if_hr_portal_maintenance()
        if hr_maintenance_response is not False:
            return hr_maintenance_response

        if self._is_hr_manager():
            return request.redirect('/my/hr/admin')

        employee = self._get_employee()
        maintenance_response = self._redirect_if_hr_portal_maintenance()
        if maintenance_response is not False:
            return maintenance_response
        redirect_response = self._redirect_if_no_employee(employee)
        if redirect_response:
            return redirect_response

        fy_value = (kwargs.get('fy') or '').strip()
        month_value = (kwargs.get('month') or '').strip()

        attendance_period = self._get_selected_payroll_period(
            fy_value=fy_value,
            month_value=month_value,
        )
        selected_month = attendance_period['selected_month']
        month_navigation = self._get_month_navigation(selected_month)

        attendance_matrix_row = self._build_attendance_matrix(employee, selected_month)
        attendance_days = attendance_matrix_row['days']
        attendance_display = self._get_attendance_logs(employee)
        attendance_fy_summary = self._get_employee_attendance_fy_summary(employee, selected_month)

        values = self._prepare_portal_values(employee, {
            'attendance_matrix_row': attendance_matrix_row,
            'attendance_days': attendance_days,
            'attendances': attendance_display,
            'month_display': month_navigation.get('month_display', ''),
            'previous_month_value': month_navigation.get('previous_month_value', ''),
            'next_month_value': month_navigation.get('next_month_value', ''),
            'selected_month_value': selected_month.strftime('%Y-%m'),
            'attendance_fy_summary': attendance_fy_summary,
            'fiscal_year_label': f"FY-{attendance_period['selected_fy_code']}",
            'fiscal_year_code': attendance_period['selected_fy_code'],
            'fiscal_year_range_label': attendance_period['selected_fy_range_label'],
            'fiscal_year_month_options': attendance_period['fiscal_year_month_options'],
            'fiscal_year_options': attendance_period['fiscal_year_options'],
            'selected_fiscal_year_value': attendance_period['selected_fy_value'],
            'selected_fiscal_year_code': attendance_period['selected_fy_code'],
        })

        return request.render('hr_employee_portal.hr_employee_attendance_page', values)
    # ---------------------------------------------------------
    # EMPLOYEE: Daily Work Report
    # ---------------------------------------------------------
    @http.route('/my/hr/daily-work-report', type='http', auth='user', website=True)
    def my_hr_daily_work_report(self, **kwargs):
        hr_maintenance_response = self._redirect_if_hr_portal_maintenance()
        if hr_maintenance_response is not False:
            return hr_maintenance_response

        if self._is_hr_manager():
            return request.redirect('/my/hr/admin')

        employee = self._get_employee()
        maintenance_response = self._redirect_if_hr_portal_maintenance()
        if maintenance_response is not False:
            return maintenance_response
        redirect_response = self._redirect_if_no_employee(employee)
        if redirect_response:
            return redirect_response

        today = fields.Date.context_today(request.env.user)
        yesterday = today - timedelta(days=1)
        existing_report = False
        yesterday_report = False
        can_submit_yesterday_late = False
        temp_late_accesses = self._get_active_late_report_accesses(employee)

        if 'hr.daily.work.report' in request.env:
            report_env = request.env['hr.daily.work.report'].sudo()

            existing_report = report_env.search([
                ('employee_id', '=', employee.id),
                ('report_date', '=', today),
            ], limit=1)

            yesterday_report = report_env.search([
                ('employee_id', '=', employee.id),
                ('report_date', '=', yesterday),
            ], limit=1)

            can_submit_yesterday_late = not bool(yesterday_report)

        report_status = (kwargs.get('report_status') or '').strip()
        report_message = (kwargs.get('report_message') or '').strip()
        success = True if report_status == 'success' else False

        values = self._prepare_portal_values(employee, {
            'today': today,
            'yesterday': yesterday,
            'existing_report': existing_report,
            'yesterday_report': yesterday_report,
            'can_submit_yesterday_late': can_submit_yesterday_late,
            'temp_late_accesses': temp_late_accesses,
            'success': success,
            'report_status': report_status,
            'report_message': report_message,
        })
        return request.render('hr_employee_portal.hr_employee_daily_work_report_page', values)

    @http.route('/my/hr/daily-work-report/submit', type='http', auth='user', methods=['POST'], website=True)
    def my_hr_daily_work_report_submit(self, **post):
        hr_maintenance_response = self._redirect_if_hr_portal_maintenance()
        if hr_maintenance_response is not False:
            return hr_maintenance_response

        if self._is_hr_manager():
            return request.redirect('/my/hr/admin')

        employee = self._get_employee()
        maintenance_response = self._redirect_if_hr_portal_maintenance()
        if maintenance_response is not False:
            return maintenance_response
        if not employee:
            redirect_url = self._build_redirect_url('/my/hr', {
                'report_status': 'error',
                'report_message': 'Employee record not found.',
            })
            return request.redirect(redirect_url)

        if 'hr.daily.work.report' not in request.env:
            redirect_url = self._build_redirect_url('/my/hr/daily-work-report', {
                'report_status': 'error',
                'report_message': 'Daily Work Report module is not available.',
            })
            return request.redirect(redirect_url)

        work_mode = (post.get('work_mode') or '').strip()
        task_report = (post.get('task_report') or '').strip()

        today = fields.Date.context_today(request.env.user)
        yesterday = today - timedelta(days=1)
        requested_report_date = self._parse_portal_date(post.get('report_date')) or today

        temp_late_access = False

        if requested_report_date == today:
            report_date = today
            late_submission_source = False
        elif requested_report_date == yesterday:
            report_date = yesterday
            late_submission_source = 'auto_next_day'
        else:
            temp_late_access = self._get_active_late_report_access(employee, requested_report_date)
            if not temp_late_access:
                redirect_url = self._build_redirect_url('/my/hr/daily-work-report', {
                    'report_status': 'error',
                    'report_message': 'Temporary access is not available or has expired for this missed report date.',
                })
                return request.redirect(redirect_url)

            report_date = requested_report_date
            late_submission_source = 'admin_temp_access'

        valid_work_modes = self._get_valid_work_modes()
        if work_mode not in valid_work_modes:
            redirect_url = self._build_redirect_url('/my/hr/daily-work-report', {
                'report_status': 'error',
                'report_message': 'Please select a valid work mode.',
            })
            return request.redirect(redirect_url)

        if not task_report:
            redirect_url = self._build_redirect_url('/my/hr/daily-work-report', {
                'report_status': 'error',
                'report_message': 'Task Report is required.',
            })
            return request.redirect(redirect_url)

        existing_report = request.env['hr.daily.work.report'].sudo().search([
            ('employee_id', '=', employee.id),
            ('report_date', '=', report_date),
        ], limit=1)
        if existing_report:
            redirect_url = self._build_redirect_url('/my/hr/daily-work-report', {
                'report_status': 'error',
                'report_message': 'You have already submitted your report for this date.',
            })
            return request.redirect(redirect_url)

        try:
            new_report = request.env['hr.daily.work.report'].sudo().with_context(
                enforce_employee_portal_rule=True
            ).create({
                'employee_id': employee.id,
                'report_date': report_date,
                'work_mode': work_mode,
                'task_report': task_report,
                'state': 'submitted',
                'submitted_at': fields.Datetime.now(),
                'late_submission_source': late_submission_source,
            })

            if late_submission_source == 'admin_temp_access':
                used_accesses = request.env['hr.late.report.access'].sudo().search([
                    ('employee_id', '=', employee.id),
                    ('report_date', '=', report_date),
                    ('state', '=', 'active'),
                ])
                if used_accesses:
                    used_accesses.write({
                        'state': 'used',
                        'used_report_id': new_report.id,
                    })
        except ValidationError as error:
            redirect_url = self._build_redirect_url('/my/hr/daily-work-report', {
                'report_status': 'error',
                'report_message': str(error),
            })
            return request.redirect(redirect_url)
        except Exception:
            redirect_url = self._build_redirect_url('/my/hr/daily-work-report', {
                'report_status': 'error',
                'report_message': 'Report could not be submitted. Please contact HR.',
            })
            return request.redirect(redirect_url)

        success_message = (
            'Late work report submitted successfully for %s.' % report_date
            if late_submission_source
            else 'Daily work report submitted successfully.'
        )

        redirect_url = self._build_redirect_url('/my/hr/daily-work-report', {
            'report_status': 'success',
            'report_message': success_message,
        })
        return request.redirect(redirect_url)
    @http.route('/my/hr/admin/performance', type='http', auth='user', website=True)
    def my_hr_admin_performance(self, **kwargs):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        self._expire_late_report_accesses()

        performance_env = request.env['hr.daily.performance.plan'].sudo()
        employee_env = request.env['hr.employee'].sudo()

        employee_id_raw = (kwargs.get('employee_id') or '').strip()
        selected_employee_id = False
        if employee_id_raw:
            try:
                selected_employee_id = int(employee_id_raw)
            except (TypeError, ValueError):
                selected_employee_id = False

        date_from = self._parse_portal_date(kwargs.get('date_from'))
        date_to = self._parse_portal_date(kwargs.get('date_to'))

        domain = self._get_performance_search_domain(
            employee_id=selected_employee_id,
            date_from=date_from,
            date_to=date_to,
    )

        try:
            page = int(kwargs.get('page') or 1)
        except (TypeError, ValueError):
            page = 1

        page = max(page, 1)
        per_page = 20
        offset = (page - 1) * per_page

        performance_total = performance_env.search_count(domain)

        performance_records = performance_env.search(
            domain,
            order='plan_date desc, id desc',
            limit=per_page,
            offset=offset,
    )

        page_count = max((performance_total + per_page - 1) // per_page, 1)

        performance_start = offset + 1 if performance_total else 0
        performance_end = min(offset + per_page, performance_total)

        base_params = {
            'employee_id': selected_employee_id,
            'date_from': date_from.strftime('%Y-%m-%d') if date_from else '',
            'date_to': date_to.strftime('%Y-%m-%d') if date_to else '',
    }

        previous_page_url = False
        next_page_url = False

        if page > 1:
            previous_params = dict(base_params)
            previous_params['page'] = page - 1
            previous_page_url = self._build_redirect_url('/my/hr/admin/performance', previous_params)

        if page < page_count:
            next_params = dict(base_params)
            next_params['page'] = page + 1
            next_page_url = self._build_redirect_url('/my/hr/admin/performance', next_params)

        employees = employee_env.search([], order='name asc')

        return request.render(
            'hr_employee_portal.hr_admin_performance_page',
            self._prepare_portal_values(None, {
                'performance_records': performance_records,
                'employees': employees,
                'selected_employee_id': selected_employee_id,
                'date_from': date_from.strftime('%Y-%m-%d') if date_from else '',
                'date_to': date_to.strftime('%Y-%m-%d') if date_to else '',
                'performance_total': performance_total,
                'performance_page': page,
                'performance_page_count': page_count,
                'performance_per_page': per_page,
                'performance_offset': offset,
                'performance_start': performance_start,
                'performance_end': performance_end,
                'previous_page_url': previous_page_url,
                'next_page_url': next_page_url,
                'performance_status': (kwargs.get('performance_status') or '').strip(),
                'performance_message': (kwargs.get('performance_message') or '').strip(),
        })
    )
    @http.route('/my/hr/admin/performance/update', type='http', auth='user', methods=['POST'], website=True)
    def my_hr_admin_performance_update(self, **post):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        if 'hr.daily.performance.plan' not in request.env:
            return request.redirect(self._build_redirect_url('/my/hr/admin/performance', {
                'performance_status': 'error',
                'performance_message': 'Performance module is not available.',
            }))

        record_id_raw = (post.get('record_id') or '').strip()
        try:
            record_id = int(record_id_raw)
        except (TypeError, ValueError):
            return request.redirect(self._build_redirect_url('/my/hr/admin/performance', {
                'performance_status': 'error',
                'performance_message': 'Invalid performance record selected.',
            }))

        record = request.env['hr.daily.performance.plan'].sudo().browse(record_id).exists()
        if not record:
            return request.redirect(self._build_redirect_url('/my/hr/admin/performance', {
                'performance_status': 'error',
                'performance_message': 'Performance record not found.',
            }))

        employee_id_raw = (post.get('employee_id') or '').strip()
        try:
            employee_id = int(employee_id_raw)
        except (TypeError, ValueError):
            employee_id = False

        plan_date = self._parse_portal_date(post.get('plan_date'))
        task_id = (post.get('task_id') or '').strip()
        project_name = (post.get('project_name') or '').strip()
        task_description = (post.get('task_description') or '').strip()
        priority_level = (post.get('priority_level') or '').strip()
        status = (post.get('status') or '').strip()
        completion_raw = (post.get('completion_percent') or '').strip()
        supervisor_remarks = (post.get('supervisor_remarks') or '').strip()

        try:
            completion_percent = float(completion_raw or 0.0)
        except (TypeError, ValueError):
            return request.redirect(self._build_redirect_url('/my/hr/admin/performance', {
                'performance_status': 'error',
                'performance_message': 'Completion percent must be a valid number.',
            }))

        if not employee_id:
            return request.redirect(self._build_redirect_url('/my/hr/admin/performance', {
                'performance_status': 'error',
                'performance_message': 'Please select a valid employee.',
            }))

        if not plan_date:
            return request.redirect(self._build_redirect_url('/my/hr/admin/performance', {
                'performance_status': 'error',
                'performance_message': 'Please enter a valid date.',
            }))

        if not task_id or not project_name or not task_description:
            return request.redirect(self._build_redirect_url('/my/hr/admin/performance', {
                'performance_status': 'error',
                'performance_message': 'Task ID, Project, and Task Description are required.',
            }))

        if priority_level not in self._get_valid_performance_priorities():
            return request.redirect(self._build_redirect_url('/my/hr/admin/performance', {
                'performance_status': 'error',
                'performance_message': 'Please select a valid priority level.',
            }))

        if status not in self._get_valid_performance_statuses():
            return request.redirect(self._build_redirect_url('/my/hr/admin/performance', {
                'performance_status': 'error',
                'performance_message': 'Please select a valid status.',
            }))

        try:
            record.write({
                'employee_id': employee_id,
                'plan_date': plan_date,
                'task_id': task_id,
                'project_name': project_name,
                'task_description': task_description,
                'priority_level': priority_level,
                'status': status,
                'completion_percent': completion_percent,
                'supervisor_remarks': supervisor_remarks,
            })
        except ValidationError as error:
            return request.redirect(self._build_redirect_url('/my/hr/admin/performance', {
                'performance_status': 'error',
                'performance_message': str(error),
            }))
        except Exception:
            return request.redirect(self._build_redirect_url('/my/hr/admin/performance', {
                'performance_status': 'error',
                'performance_message': 'Performance record could not be updated.',
            }))

        return request.redirect(self._build_redirect_url('/my/hr/admin/performance', {
            'performance_status': 'success',
            'performance_message': 'Performance record updated successfully.',
        }))
    @http.route('/my/hr/admin/performance/export', type='http', auth='user', website=True)
    def my_hr_admin_performance_export(self, **kwargs):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        employee_id_raw = (kwargs.get('employee_id') or '').strip()
        selected_employee_id = False

        if employee_id_raw:
            try:
                selected_employee_id = int(employee_id_raw)
            except (TypeError, ValueError):
                selected_employee_id = False

        date_from = self._parse_portal_date(kwargs.get('date_from'))
        date_to = self._parse_portal_date(kwargs.get('date_to'))

        domain = self._get_performance_search_domain(
            employee_id=selected_employee_id,
            date_from=date_from,
            date_to=date_to,
        )

        records = request.env['hr.daily.performance.plan'].sudo().search(
            domain,
            order='plan_date desc, id desc',
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Performance'

        sheet.append([
            '#',
            'Employee',
            'Employee Code',
            'Date',
            'Task ID',
            'Project',
            'Task Description',
            'Priority',
            'Status',
            'Completion %',
            'Supervisor Remarks',
        ])

        for index, rec in enumerate(records, start=1):
            sheet.append([
                index,
                rec.employee_id.name or '',
                rec.employee_id.employee_code or '',
                rec.plan_date.strftime('%Y-%m-%d') if rec.plan_date else '',
                rec.task_id or '',
                rec.project_name or '',
                rec.task_description or '',
                rec.priority_level or '',
                rec.status or '',
                rec.completion_percent or 0.0,
                rec.supervisor_remarks or '',
            ])

        return self._export_workbook_response(workbook, 'performance_records.xlsx')

    # ---------------------------------------------------------
    # EMPLOYEE: Documents, Payroll, Leaves
    # ---------------------------------------------------------
    @http.route('/my/hr/documents', type='http', auth='user', website=True)
    def my_hr_documents(self, **kwargs):
        hr_maintenance_response = self._redirect_if_hr_portal_maintenance()
        if hr_maintenance_response is not False:
            return hr_maintenance_response

        if self._is_hr_manager():
            return request.redirect('/my/hr/admin')

        return request.render(
            'hr_employee_portal.hr_employee_documents_page',
            self._prepare_portal_values(self._get_employee())
        )

    @http.route('/my/hr/payroll', type='http', auth='user', website=True)
    def my_hr_payroll(self, **kwargs):
        hr_maintenance_response = self._redirect_if_hr_portal_maintenance()
        if hr_maintenance_response is not False:
            return hr_maintenance_response

        if self._is_hr_manager():
            return request.redirect('/my/hr/admin')

        employee = self._get_employee()
        maintenance_response = self._redirect_if_hr_portal_maintenance()
        if maintenance_response is not False:
            return maintenance_response
        fy_value = kwargs.get('fy')
        month_value = kwargs.get('month')

        payroll_period = self._get_selected_payroll_period(fy_value=fy_value, month_value=month_value)
        selected_month = payroll_period['selected_month']
        month_navigation = self._get_month_navigation(selected_month)

        payroll_data = False
        if employee:
            payroll_data = self._get_employee_payroll_data(employee, selected_month)

        return request.render(
            'hr_employee_portal.hr_employee_payslips_page',
            self._prepare_portal_values(employee, {
                'payroll_data': payroll_data,
                'month_display': month_navigation.get('month_display', ''),
                'previous_month_value': month_navigation.get('previous_month_value', ''),
                'next_month_value': month_navigation.get('next_month_value', ''),
                'selected_month_value': selected_month.strftime('%Y-%m'),
                'fiscal_year_label': f"FY-{payroll_period['selected_fy_code']}",
                'fiscal_year_code': payroll_period['selected_fy_code'],
                'fiscal_year_range_label': payroll_period['selected_fy_range_label'],
                'fiscal_year_month_options': payroll_period['fiscal_year_month_options'],
                'fiscal_year_options': payroll_period['fiscal_year_options'],
                'selected_fiscal_year_value': payroll_period['selected_fy_value'],
                'selected_fiscal_year_code': payroll_period['selected_fy_code'],
                'attendance_status': (kwargs.get('attendance_status') or '').strip(),
                'attendance_message': (kwargs.get('attendance_message') or '').strip(),
                'payroll_tax_section_id': 'employee-payroll-tax-section',
                'payroll_tax_rate_section_id': 'employee-payroll-tax-rate-section',
            })
        )

    @http.route('/my/hr/payslips', type='http', auth='user', website=True)
    def my_hr_payslips(self, **kwargs):
        hr_maintenance_response = self._redirect_if_hr_portal_maintenance()
        if hr_maintenance_response is not False:
            return hr_maintenance_response

        if self._is_hr_manager():
            return request.redirect('/my/hr/admin')
        return request.redirect('/my/hr/payroll')

    @http.route('/my/hr/leaves', type='http', auth='user', website=True)
    def my_hr_leaves(self, **kwargs):
        hr_maintenance_response = self._redirect_if_hr_portal_maintenance()
        if hr_maintenance_response is not False:
            return hr_maintenance_response

        if self._is_hr_manager():
            return request.redirect('/my/hr/admin')

        employee = self._get_employee()
        maintenance_response = self._redirect_if_hr_portal_maintenance()
        if maintenance_response is not False:
            return maintenance_response
        hr_requests = request.env['hr.employee.portal.request'].sudo()

        try:
            request_page = int(kwargs.get('request_page') or 1)
        except (TypeError, ValueError):
            request_page = 1

        request_page = max(request_page, 1)
        per_page = 8
        offset = (request_page - 1) * per_page

        request_total = 0
        request_page_count = 1
        request_start = 0
        request_end = 0
        previous_request_page_url = False
        next_request_page_url = False

        if employee and 'hr.employee.portal.request' in request.env:
            domain = [('employee_id', '=', employee.id)]

            request_total = hr_requests.search_count(domain)
            hr_requests = hr_requests.search(
                domain,
                order='submitted_at desc, id desc',
                limit=per_page,
                offset=offset,
            )

            request_page_count = max((request_total + per_page - 1) // per_page, 1)
            request_start = offset + 1 if request_total else 0
            request_end = min(offset + per_page, request_total)

            if request_page > 1:
                previous_request_page_url = self._build_redirect_url('/my/hr/leaves', {
                    'request_page': request_page - 1,
                })

            if request_page < request_page_count:
                next_request_page_url = self._build_redirect_url('/my/hr/leaves', {
                    'request_page': request_page + 1,
                })

        return request.render(
            'hr_employee_portal.hr_employee_leaves_page',
            self._prepare_portal_values(employee, {
                'hr_requests': hr_requests,
                'request_total': request_total,
                'request_page': request_page,
                'request_page_count': request_page_count,
                'request_start': request_start,
                'request_end': request_end,
                'previous_request_page_url': previous_request_page_url,
                'next_request_page_url': next_request_page_url,
            })
        )

        # =========================================================
    # ADMIN: Secure Manager Routes
    # =========================================================
    @http.route('/my/hr/admin', type='http', auth='user', website=True)
    def my_hr_admin_dashboard(self, **kwargs):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        hr_employee_env = request.env['hr.employee'].sudo().with_context(active_test=False)
        hr_task_report_env = request.env['hr.daily.work.report'].sudo()
        attendance_register_env = request.env['hr.attendance.register.line'].sudo()

        all_employees = hr_employee_env.search([], order='name asc')

        today = fields.Date.context_today(request.env.user)
        month_start = today.replace(day=1)
        total_days = calendar.monthrange(today.year, today.month)[1]
        month_end = today.replace(day=total_days)

        today_start = datetime.combine(today, time.min)
        next_day_start = today_start + timedelta(days=1)

        total_employees = hr_employee_env.search_count([])

        month_register_lines = attendance_register_env.search([
            ('attendance_date', '>=', month_start),
            ('attendance_date', '<=', month_end),
            ('attendance_code', 'in', ['P', 'R']),
        ])
        present_employee_count = len(set(month_register_lines.mapped('employee_id').ids))

        task_reports_submitted_today = hr_task_report_env.search_count([
            ('submitted_at', '>=', fields.Datetime.to_string(today_start)),
            ('submitted_at', '<', fields.Datetime.to_string(next_day_start)),
        ])

        hr_maintenance_enabled = self._is_hr_portal_maintenance_enabled()

        values = self._prepare_portal_values(None, {
            'employees': all_employees,
            'total_employees': total_employees,
            'present_employees': present_employee_count,
            'task_reports_submitted_today': task_reports_submitted_today,
            'hr_maintenance_enabled': hr_maintenance_enabled,
            'employee_status': (kwargs.get('employee_status') or '').strip(),
            'employee_message': (kwargs.get('employee_message') or '').strip(),
        })
        return request.render('hr_employee_portal.hr_admin_dashboard_page', values)



    @http.route('/my/hr/admin/maintenance/toggle', type='http', auth='user', methods=['POST'], website=True)
    def my_hr_admin_maintenance_toggle(self, **post):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        action = (post.get('maintenance_action') or '').strip()
        enabled = '1' if action == 'on' else '0'

        request.env['ir.config_parameter'].sudo().set_param(
            'hr_employee_portal.maintenance_mode',
            enabled
        )

        message = (
            'Employee HR portal maintenance mode enabled. Employees are temporarily blocked.'
            if enabled == '1'
            else 'Employee HR portal maintenance mode disabled. Employees can access the portal again.'
        )

        return request.redirect(self._build_redirect_url('/my/hr/admin', {
            'employee_status': 'success',
            'employee_message': message,
        }))

    @http.route('/my/hr/admin/portal-accounts', type='http', auth='user', website=True)
    def my_hr_admin_portal_accounts_page(self, **kwargs):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        values = self._prepare_portal_values(None, {
            'employee_status': (kwargs.get('employee_status') or '').strip(),
            'employee_message': (kwargs.get('employee_message') or '').strip(),
        })
        return request.render('hr_employee_portal.hr_admin_portal_accounts_page', values)

    @http.route('/my/hr/admin/employee/portal-account/save', type='http', auth='user', methods=['POST'], website=True)
    def my_hr_admin_employee_portal_account_save(self, **post):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        employee_code = (post.get('employee_code') or '').strip()
        employee_name = (post.get('employee_name') or '').strip()
        work_email = (post.get('work_email') or '').strip().lower()
        temp_password = (post.get('temp_password') or '').strip()
        department_name = (post.get('department_name') or '').strip()
        job_title = (post.get('job_title') or '').strip()

        if not employee_code or not employee_name or not work_email or not temp_password:
            return request.redirect(self._build_redirect_url('/my/hr/admin', {
                'employee_status': 'error',
                'employee_message': 'Employee code, name, official email, and temporary password are required.',
            }))

        if '@' not in work_email or '.' not in work_email:
            return request.redirect(self._build_redirect_url('/my/hr/admin', {
                'employee_status': 'error',
                'employee_message': 'Please enter a valid official email.',
            }))

        Employee = request.env['hr.employee'].sudo().with_context(active_test=False)
        User = request.env['res.users'].sudo().with_context(
            active_test=False,
            no_reset_password=True,
            tracking_disable=True,
            mail_notrack=True,
            mail_create_nosubscribe=True,
            mail_notify_noemail=True,
        )
        Department = request.env['hr.department'].sudo()
        Job = request.env['hr.job'].sudo()

        existing_code_employee = Employee.search([('employee_code', '=', employee_code)], limit=1)
        existing_email_user = User.search(['|', ('login', '=', work_email), ('email', '=', work_email)], limit=1)

        department = False
        if department_name:
            department = Department.search([('name', '=ilike', department_name)], limit=1)
            if not department:
                department = Department.create({'name': department_name})

        job = False
        if job_title:
            job = Job.search([('name', '=ilike', job_title)], limit=1)
            if not job:
                job = Job.create({'name': job_title})

        employee_vals = {
            'name': employee_name,
            'employee_code': employee_code,
            'work_email': work_email,
            'active': True,
        }
        if department:
            employee_vals['department_id'] = department.id
        if job:
            employee_vals['job_id'] = job.id

        if existing_code_employee:
            employee = existing_code_employee
            employee.write(employee_vals)
            employee_action = 'updated'
        else:
            employee = Employee.create(employee_vals)
            employee_action = 'created'

        portal_group = request.env.ref('base.group_portal')
        company = employee.company_id or request.env.company

        user = employee.user_id or existing_email_user

        # Do not convert internal/admin users into portal users.
        if user and not user.share:
            employee.write({'user_id': user.id})
            return request.redirect(self._build_redirect_url('/my/hr/admin', {
                'employee_status': 'error',
                'employee_message': 'Employee updated, but linked user is an internal/admin user. Portal role was not changed for safety.',
            }))

        user_vals = {
            'name': employee.name,
            'login': work_email,
            'email': work_email,
            'active': True,
            'company_id': company.id,
            'company_ids': [(6, 0, [company.id])],
            'group_ids': [(6, 0, [portal_group.id])],
            'password': temp_password,
        }

        if user:
            user.write(user_vals)
            user_action = 'updated'
        else:
            user = User.create(user_vals)
            user_action = 'created'

        employee.write({'user_id': user.id})

        # Remove unwanted automatic account/security emails generated by user creation/password changes.
        blocked_subject_parts = [
            'Welcome',
            'Password Changed',
            'Login Changed',
            'Your account',
            'Security Update',
        ]
        Mail = request.env['mail.mail'].sudo()
        mails = Mail.search([('email_to', 'ilike', work_email)])
        for mail in mails:
            subject = (mail.mail_message_id.subject or '')
            if any(part.lower() in subject.lower() for part in blocked_subject_parts):
                mail.unlink()

        return request.redirect(self._build_redirect_url('/my/hr/admin', {
            'employee_status': 'success',
            'employee_message': f'Employee {employee_action} and portal account {user_action} for {employee.name}. No account email was sent.',
        }))


    @http.route('/my/hr/admin/employee/status/update', type='http', auth='user', methods=['POST'], website=True)
    def my_hr_admin_employee_status_update(self, **post):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        employee_id_raw = (post.get('employee_id') or '').strip()
        active_value = (post.get('active') or '').strip()

        try:
            employee_id = int(employee_id_raw)
        except (TypeError, ValueError):
            return request.redirect(self._build_redirect_url('/my/hr/admin', {
                'employee_status': 'error',
                'employee_message': 'Invalid employee selected.',
            }))

        employee = request.env['hr.employee'].sudo().with_context(active_test=False).browse(employee_id).exists()
        if not employee:
            return request.redirect(self._build_redirect_url('/my/hr/admin', {
                'employee_status': 'error',
                'employee_message': 'Employee not found.',
            }))

        if employee.user_id and employee.user_id.id == request.env.user.id and active_value == '0':
            return request.redirect(self._build_redirect_url('/my/hr/admin', {
                'employee_status': 'error',
                'employee_message': 'You cannot deactivate your own linked employee record from the portal.',
            }))

        if active_value not in ['0', '1']:
            return request.redirect(self._build_redirect_url('/my/hr/admin', {
                'employee_status': 'error',
                'employee_message': 'Invalid status selected.',
            }))

        employee.write({
            'active': True if active_value == '1' else False,
        })

        return request.redirect(self._build_redirect_url('/my/hr/admin', {
            'employee_status': 'success',
            'employee_message': 'Employee status updated successfully.',
        }))

    @http.route('/my/hr/admin/attendances', type='http', auth='user', website=True)
    def my_hr_admin_attendances(self, **kwargs):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        hr_employee_env = request.env['hr.employee'].sudo()
        employees = hr_employee_env.search([], order='name asc')

        selected_employee = False
        selected_employee_id = kwargs.get('employee_id')
        fy_value = kwargs.get('fy')
        month_value = kwargs.get('month')

        payroll_period = self._get_selected_payroll_period(
            fy_value=fy_value,
            month_value=month_value,
        )
        selected_month = payroll_period['selected_month']
        month_navigation = self._get_month_navigation(selected_month)
        fiscal_context = self._get_fiscal_year_context(selected_month)

        if selected_employee_id:
            try:
                selected_employee_id = int(selected_employee_id)
                selected_employee = hr_employee_env.browse(selected_employee_id).exists()
            except (TypeError, ValueError):
                selected_employee = False

        if not selected_employee and employees:
            selected_employee = employees[0]
            selected_employee_id = selected_employee.id
        elif selected_employee:
            selected_employee_id = selected_employee.id
        else:
            selected_employee_id = False

        attendance_matrix_row = False
        attendance_days = []
        attendance_display = []
        attendance_fy_summary = False

        if selected_employee:
            attendance_matrix_row = self._build_attendance_matrix(selected_employee, selected_month)
            attendance_days = attendance_matrix_row['days']
            attendance_display = self._get_attendance_logs(selected_employee)
            attendance_fy_summary = self._get_employee_attendance_fy_summary(selected_employee, selected_month)

        values = self._prepare_portal_values(None, {
            'employees': employees,
            'selected_employee': selected_employee,
            'selected_employee_id': selected_employee_id,
            'attendance_matrix_row': attendance_matrix_row,
            'attendance_days': attendance_days,
            'attendances': attendance_display,
            'attendance_fy_summary': attendance_fy_summary,

            'month_display': month_navigation.get('month_display', ''),
            'previous_month_value': month_navigation.get('previous_month_value', ''),
            'next_month_value': month_navigation.get('next_month_value', ''),
            'selected_month_value': selected_month.strftime('%Y-%m'),

            'fiscal_year_label': fiscal_context['fiscal_year_label'],
            'fiscal_year_code': fiscal_context['fiscal_year_code'],
            'fiscal_year_range_label': fiscal_context['fiscal_year_range_label'],
            'fiscal_year_month_options': payroll_period['fiscal_year_month_options'],
            'fiscal_year_options': payroll_period['fiscal_year_options'],
            'selected_fiscal_year_value': payroll_period['selected_fy_value'],
            'selected_fiscal_year_code': payroll_period['selected_fy_code'],

            'attendance_status': (kwargs.get('attendance_status') or '').strip(),
            'attendance_message': (kwargs.get('attendance_message') or '').strip(),
        })

        return request.render(
            'hr_employee_portal.hr_admin_attendances_page',
            values
        )
    @http.route('/my/hr/admin/attendances/update', type='http', auth='user', methods=['POST'], website=True)
    def my_hr_admin_attendances_update(self, **post):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        employee_id_raw = (post.get('employee_id') or '').strip()
        fy_value = (post.get('fy') or '').strip()
        month_value = (post.get('month') or '').strip()

        base_params = {
            'employee_id': employee_id_raw,
            'fy': fy_value,
            'month': month_value,
        }

        try:
            employee_id = int(employee_id_raw)
        except (TypeError, ValueError):
            base_params.update({
                'attendance_status': 'error',
                'attendance_message': 'Invalid employee selected.',
            })
            return request.redirect(self._build_redirect_url('/my/hr/admin/attendances', base_params))

        employee = request.env['hr.employee'].sudo().browse(employee_id).exists()
        if not employee:
            base_params.update({
                'attendance_status': 'error',
                'attendance_message': 'Employee not found.',
            })
            return request.redirect(self._build_redirect_url('/my/hr/admin/attendances', base_params))

        attendance_period = self._get_selected_payroll_period(
            fy_value=fy_value,
            month_value=month_value,
        )
        selected_month = attendance_period['selected_month']
        month_start = selected_month.replace(day=1)
        total_days = calendar.monthrange(month_start.year, month_start.month)[1]

        valid_codes = {'P', 'R', 'H', 'S', 'C', 'U', 'D', 'OT'}
        register_env = request.env['hr.attendance.register.line'].sudo()

        fiscal_year_label = self._get_fiscal_year_label(month_start)
        month_label = month_start.strftime('%B %Y')

        try:
            for day_number in range(1, total_days + 1):
                raw_code = (post.get(f'day_{day_number}') or '').strip().upper()
                attendance_date = month_start.replace(day=day_number)

                existing_line = register_env.search([
                    ('employee_id', '=', employee.id),
                    ('attendance_date', '=', attendance_date),
                ], limit=1)

                if raw_code in ('', '-'):
                    if existing_line:
                        existing_line.unlink()
                    continue

                if raw_code not in valid_codes:
                    raise ValidationError(f"Invalid attendance code '{raw_code}' for day {day_number}.")

                register_env.create_or_update_attendance_line(
                    employee=employee,
                    attendance_date=attendance_date,
                    attendance_code=raw_code,
                    fiscal_year_label=fiscal_year_label,
                    month_label=month_label,
                    source='manual',
                    notes='Updated manually from admin portal.',
                )

        except ValidationError as error:
            base_params.update({
                'attendance_status': 'error',
                'attendance_message': str(error),
            })
            return request.redirect(self._build_redirect_url('/my/hr/admin/attendances', base_params))
        except Exception:
            base_params.update({
                'attendance_status': 'error',
                'attendance_message': 'Attendance could not be updated.',
            })
            return request.redirect(self._build_redirect_url('/my/hr/admin/attendances', base_params))

        base_params.update({
            'attendance_status': 'success',
            'attendance_message': 'Attendance updated successfully.',
        })
        return request.redirect(self._build_redirect_url('/my/hr/admin/attendances', base_params))

    @http.route('/my/hr/admin/attendances/export', type='http', auth='user', website=True)
    def my_hr_admin_attendances_export(self, **kwargs):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        employee_id_raw = (kwargs.get('employee_id') or '').strip()
        fy_value = (kwargs.get('fy') or '').strip()
        month_value = (kwargs.get('month') or '').strip()

        try:
            employee_id = int(employee_id_raw)
        except (TypeError, ValueError):
            return request.redirect('/my/hr/admin/attendances')

        employee = request.env['hr.employee'].sudo().browse(employee_id).exists()
        if not employee:
            return request.redirect('/my/hr/admin/attendances')

        attendance_period = self._get_selected_payroll_period(
            fy_value=fy_value,
            month_value=month_value,
        )
        selected_month = attendance_period['selected_month']
        attendance_matrix_row = self._build_attendance_matrix(employee, selected_month)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Attendance'

        day_headers = [
            'Day %s' % day_data['day_number']
            for day_data in attendance_matrix_row['days']
        ]

        headers = [
            'Employee Code',
            'Employee Name',
            'Fiscal Year',
            'Month',
        ] + day_headers + [
            'P',
            'R',
            'H',
            'S',
            'C',
            'U',
            'D',
            'OT',
        ]

        sheet.append(headers)

        day_values = [
            day_data.get('code') or '-'
            for day_data in attendance_matrix_row['days']
        ]

        summary = attendance_matrix_row['summary']

        sheet.append([
            employee.employee_code or '',
            employee.name or '',
            attendance_matrix_row['fiscal_year'],
            attendance_matrix_row['month_label'],
        ] + day_values + [
            summary.get('P', 0),
            summary.get('R', 0),
            summary.get('H', 0),
            summary.get('S', 0),
            summary.get('C', 0),
            summary.get('U', 0),
            summary.get('D', 0),
            summary.get('OT', 0),
        ])

        file_name = 'attendance_%s_%s.xlsx' % (
            employee.employee_code or employee.id,
            selected_month.strftime('%Y_%m'),
        )

        return self._export_workbook_response(workbook, file_name)
    @http.route('/my/hr/admin/task-reports', type='http', auth='user', website=True)
    def my_hr_admin_task_reports(self, **kwargs):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        self._expire_late_report_accesses()

        hr_employee_env = request.env['hr.employee'].sudo()
        report_env = request.env['hr.daily.work.report'].sudo()

        employees = hr_employee_env.search([], order='name asc')

        selected_employee_id = False
        employee_id_raw = (kwargs.get('employee_id') or '').strip()
        if employee_id_raw:
            try:
                selected_employee_id = int(employee_id_raw)
            except (TypeError, ValueError):
                selected_employee_id = False

        date_from = self._parse_portal_date(kwargs.get('date_from'))
        date_to = self._parse_portal_date(kwargs.get('date_to'))

        domain = self._get_task_report_search_domain(
            employee_id=selected_employee_id,
            date_from=date_from,
            date_to=date_to,
    )

        try:
            page = int(kwargs.get('page') or 1)
        except (TypeError, ValueError):
            page = 1

        page = max(page, 1)
        per_page = 20
        offset = (page - 1) * per_page

        task_report_total = report_env.search_count(domain)

        task_reports = report_env.search(
            domain,
            order='report_date desc, submitted_at desc, id desc',
            limit=per_page,
            offset=offset,
    )

        page_count = max((task_report_total + per_page - 1) // per_page, 1)

        task_report_start = offset + 1 if task_report_total else 0
        task_report_end = min(offset + per_page, task_report_total)

        base_params = {
            'employee_id': selected_employee_id,
            'date_from': date_from.strftime('%Y-%m-%d') if date_from else '',
            'date_to': date_to.strftime('%Y-%m-%d') if date_to else '',
    }

        previous_page_url = False
        next_page_url = False

        if page > 1:
            previous_params = dict(base_params)
            previous_params['page'] = page - 1
            previous_page_url = self._build_redirect_url('/my/hr/admin/task-reports', previous_params)

        if page < page_count:
            next_params = dict(base_params)
            next_params['page'] = page + 1
            next_page_url = self._build_redirect_url('/my/hr/admin/task-reports', next_params)

        late_accesses = False
        late_access_employee_id_raw = (kwargs.get('late_access_employee_id') or '').strip()
        late_access_report_date = self._parse_portal_date(kwargs.get('late_access_report_date'))
        late_access_state = (kwargs.get('late_access_state') or '').strip()

        selected_late_access_employee_id = False
        if late_access_employee_id_raw:
            try:
                selected_late_access_employee_id = int(late_access_employee_id_raw)
            except (TypeError, ValueError):
                selected_late_access_employee_id = False

        if late_access_state not in ['active', 'used', 'expired', 'revoked']:
            late_access_state = ''

        late_access_filter_applied = bool(
            selected_late_access_employee_id or late_access_report_date or late_access_state
        )

        if 'hr.late.report.access' in request.env and late_access_filter_applied:
            late_access_domain = []

            if selected_late_access_employee_id:
                late_access_domain.append(('employee_id', '=', selected_late_access_employee_id))

            if late_access_report_date:
                late_access_domain.append(('report_date', '=', late_access_report_date))

            if late_access_state:
                late_access_domain.append(('state', '=', late_access_state))

            late_accesses = request.env['hr.late.report.access'].sudo().search(
                late_access_domain,
                order='create_date desc, id desc',
                limit=50
            )

        values = self._prepare_portal_values(None, {
            'employees': employees,
            'late_accesses': late_accesses,
            'late_access_filter_applied': late_access_filter_applied,
            'selected_late_access_employee_id': selected_late_access_employee_id,
            'late_access_report_date': late_access_report_date.strftime('%Y-%m-%d') if late_access_report_date else '',
            'late_access_state': late_access_state,
            'selected_employee_id': selected_employee_id,
            'date_from': date_from.strftime('%Y-%m-%d') if date_from else '',
            'date_to': date_to.strftime('%Y-%m-%d') if date_to else '',
            'task_reports': task_reports,
            'task_report_total': task_report_total,
            'task_report_page': page,
            'task_report_page_count': page_count,
            'task_report_per_page': per_page,
            'task_report_offset': offset,
            'task_report_start': task_report_start,
            'task_report_end': task_report_end,
            'previous_page_url': previous_page_url,
            'next_page_url': next_page_url,
            'report_status': (kwargs.get('report_status') or '').strip(),
            'report_message': (kwargs.get('report_message') or '').strip(),
    })

        return request.render('hr_employee_portal.hr_admin_task_reports_page', values)



    @http.route('/my/hr/admin/late-access', type='http', auth='user', website=True)
    def my_hr_admin_late_access_page(self, **kwargs):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        self._expire_late_report_accesses()

        employees = request.env['hr.employee'].sudo().search([], order='name asc')

        employee_id_raw = (kwargs.get('employee_id') or '').strip()
        report_date = self._parse_portal_date(kwargs.get('report_date'))
        state = (kwargs.get('state') or '').strip()

        selected_employee_id = False
        if employee_id_raw:
            try:
                selected_employee_id = int(employee_id_raw)
            except (TypeError, ValueError):
                selected_employee_id = False

        if state not in ['active', 'used', 'expired', 'revoked']:
            state = ''

        filter_applied = bool(selected_employee_id or report_date or state)
        late_accesses = False

        if 'hr.late.report.access' in request.env and filter_applied:
            domain = []

            if selected_employee_id:
                domain.append(('employee_id', '=', selected_employee_id))

            if report_date:
                domain.append(('report_date', '=', report_date))

            if state:
                domain.append(('state', '=', state))

            late_accesses = request.env['hr.late.report.access'].sudo().search(
                domain,
                order='create_date desc, id desc',
                limit=50,
            )

        values = self._prepare_portal_values(None, {
            'employees': employees,
            'late_accesses': late_accesses,
            'filter_applied': filter_applied,
            'selected_employee_id': selected_employee_id,
            'report_date': report_date.strftime('%Y-%m-%d') if report_date else '',
            'state': state,
            'report_status': (kwargs.get('report_status') or '').strip(),
            'report_message': (kwargs.get('report_message') or '').strip(),
        })
        return request.render('hr_employee_portal.hr_admin_late_access_page', values)

    @http.route('/my/hr/admin/task-reports/late-access/grant', type='http', auth='user', methods=['POST'], website=True)
    def my_hr_admin_late_access_grant(self, **post):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        employee_id_raw = (post.get('employee_id') or '').strip()
        report_date = self._parse_portal_date(post.get('report_date'))
        expires_at_local = self._parse_portal_datetime_local(post.get('expires_at'))
        notes = (post.get('notes') or '').strip()

        expires_at = False
        if expires_at_local:
            user_tz_name = request.env.user.tz or 'Asia/Karachi'
            if user_tz_name == 'UTC':
                user_tz_name = 'Asia/Karachi'
            user_tz = pytz.timezone(user_tz_name)
            localized_expiry = user_tz.localize(expires_at_local)
            expires_at = localized_expiry.astimezone(pytz.UTC).replace(tzinfo=None)

        try:
            employee_id = int(employee_id_raw)
        except Exception:
            employee_id = False

        employee = request.env['hr.employee'].sudo().browse(employee_id).exists() if employee_id else False

        if not employee or not report_date or not expires_at:
            return request.redirect(self._build_redirect_url('/my/hr/admin/late-access', {
                'report_status': 'error',
                'report_message': 'Please select employee, missed date, and expiry time.',
            }))

        today = fields.Date.context_today(request.env.user)
        yesterday = today - timedelta(days=1)
        if report_date >= yesterday:
            return request.redirect(self._build_redirect_url('/my/hr/admin/late-access', {
                'report_status': 'error',
                'report_message': 'Temporary access can only be granted after the automatic next-day late window has passed.',
            }))

        if expires_at <= fields.Datetime.now():
            return request.redirect(self._build_redirect_url('/my/hr/admin/late-access', {
                'report_status': 'error',
                'report_message': 'Expiry time must be in the future.',
            }))

        existing_report = request.env['hr.daily.work.report'].sudo().search([
            ('employee_id', '=', employee.id),
            ('report_date', '=', report_date),
        ], limit=1)

        if existing_report:
            return request.redirect(self._build_redirect_url('/my/hr/admin/late-access', {
                'report_status': 'error',
                'report_message': 'This employee already has a report for the selected date.',
            }))

        access_env = request.env['hr.late.report.access'].sudo()
        existing_access = access_env.search([
            ('employee_id', '=', employee.id),
            ('report_date', '=', report_date),
            ('state', '=', 'active'),
        ], limit=1)

        vals = {
            'employee_id': employee.id,
            'report_date': report_date,
            'expires_at': fields.Datetime.to_string(expires_at),
            'notes': notes,
            'state': 'active',
        }

        if existing_access:
            existing_access.write(vals)
            message = 'Temporary access updated for %s on %s.' % (employee.name, report_date)
        else:
            access_env.create(vals)
            message = 'Temporary access granted for %s on %s.' % (employee.name, report_date)

        return request.redirect(self._build_redirect_url('/my/hr/admin/late-access', {
            'report_status': 'success',
            'report_message': message,
        }))

    @http.route('/my/hr/admin/late-access/late-access/revoke', type='http', auth='user', methods=['POST'], website=True)
    def my_hr_admin_late_access_revoke(self, **post):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        access_id_raw = (post.get('access_id') or '').strip()

        try:
            access_id = int(access_id_raw)
        except Exception:
            access_id = False

        access = request.env['hr.late.report.access'].sudo().browse(access_id).exists() if access_id else False

        if access and access.state == 'active':
            access.write({'state': 'revoked'})
            status = 'success'
            message = 'Temporary access revoked.'
        else:
            status = 'error'
            message = 'Temporary access not found or already used.'

        return request.redirect(self._build_redirect_url('/my/hr/admin/late-access', {
            'report_status': status,
            'report_message': message,
        }))

    @http.route('/my/hr/admin/late-access/update', type='http', auth='user', methods=['POST'], website=True)
    def my_hr_admin_task_reports_update(self, **post):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        if 'hr.daily.work.report' not in request.env:
            redirect_url = self._build_redirect_url('/my/hr/admin/task-reports', {
                'report_status': 'error',
                'report_message': 'Daily Work Report module is not available.',
            })
            return request.redirect(redirect_url)

        report_id_raw = (post.get('report_id') or '').strip()
        try:
            report_id = int(report_id_raw)
        except (TypeError, ValueError):
            redirect_url = self._build_redirect_url('/my/hr/admin/task-reports', {
                'report_status': 'error',
                'report_message': 'Invalid report selected.',
            })
            return request.redirect(redirect_url)

        report = request.env['hr.daily.work.report'].sudo().browse(report_id).exists()
        if not report:
            redirect_url = self._build_redirect_url('/my/hr/admin/task-reports', {
                'report_status': 'error',
                'report_message': 'Task report not found.',
            })
            return request.redirect(redirect_url)

        report_date = self._parse_portal_date(post.get('report_date'))
        submitted_at_naive = self._parse_portal_datetime_local(post.get('submitted_at'))
        work_mode = (post.get('work_mode') or '').strip()
        task_report = (post.get('task_report') or '').strip()
        remarks = (post.get('remarks') or '').strip()
        state = (post.get('state') or '').strip()

        if not report_date:
            return request.redirect(self._build_redirect_url('/my/hr/admin/task-reports', {
                'report_status': 'error',
                'report_message': 'Please enter a valid report date.',
            }))

        if not submitted_at_naive:
            return request.redirect(self._build_redirect_url('/my/hr/admin/task-reports', {
                'report_status': 'error',
                'report_message': 'Please enter a valid submitted time.',
            }))

        if work_mode not in self._get_valid_work_modes():
            return request.redirect(self._build_redirect_url('/my/hr/admin/task-reports', {
                'report_status': 'error',
                'report_message': 'Please select a valid work mode.',
            }))

        if not task_report:
            return request.redirect(self._build_redirect_url('/my/hr/admin/task-reports', {
                'report_status': 'error',
                'report_message': 'Task report is required.',
            }))

        if state not in ['draft', 'submitted', 'approved', 'rejected']:
            return request.redirect(self._build_redirect_url('/my/hr/admin/task-reports', {
                'report_status': 'error',
                'report_message': 'Please select a valid state.',
            }))

        try:
            report.write({
                'report_date': report_date,
                'submitted_at': fields.Datetime.to_string(submitted_at_naive),
                'work_mode': work_mode,
                'task_report': task_report,
                'remarks': remarks,
                'state': state,
            })
        except ValidationError as error:
            return request.redirect(self._build_redirect_url('/my/hr/admin/task-reports', {
                'report_status': 'error',
                'report_message': str(error),
            }))
        except Exception:
            return request.redirect(self._build_redirect_url('/my/hr/admin/task-reports', {
                'report_status': 'error',
                'report_message': 'Task report could not be updated.',
            }))

        return request.redirect(self._build_redirect_url('/my/hr/admin/task-reports', {
            'report_status': 'success',
            'report_message': 'Task report updated successfully.',
        }))
    @http.route('/my/hr/admin/documents', type='http', auth='user', website=True)
    def my_hr_admin_documents(self, **kwargs):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')
        return request.render('hr_employee_portal.hr_admin_documents_page', self._prepare_portal_values(None))
    @http.route('/my/hr/admin/task-reports/export', type='http', auth='user', website=True)
    def my_hr_admin_task_reports_export(self, **kwargs):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        employee_id_raw = (kwargs.get('employee_id') or '').strip()
        selected_employee_id = False

        if employee_id_raw:
            try:
                selected_employee_id = int(employee_id_raw)
            except (TypeError, ValueError):
                selected_employee_id = False

        date_from = self._parse_portal_date(kwargs.get('date_from'))
        date_to = self._parse_portal_date(kwargs.get('date_to'))

        domain = self._get_task_report_search_domain(
            employee_id=selected_employee_id,
            date_from=date_from,
            date_to=date_to,
        )

        reports = request.env['hr.daily.work.report'].sudo().search(
            domain,
            order='report_date desc, submitted_at desc, id desc',
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Task Reports'

        sheet.append([
            '#',
            'Employee',
            'Employee Code',
            'Report Date',
            'Submitted At',
            'Work Mode',
            'Status',
            'Task Report',
            'Remarks',
        ])

        for index, report in enumerate(reports, start=1):
            if report.work_mode == 'office':
                work_mode = 'Work from Office'
            elif report.work_mode == 'wfh':
                work_mode = 'Work from Home'
            elif report.work_mode == 'field':
                work_mode = 'Field Work'
            elif report.work_mode == 'leave':
                work_mode = 'Leave'
            else:
                work_mode = report.work_mode or ''

            sheet.append([
                index,
                report.employee_id.name or '',
                report.employee_id.employee_code or '',
                report.report_date.strftime('%Y-%m-%d') if report.report_date else '',
                report.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if report.submitted_at else '',
                work_mode,
                report.state or '',
                report.task_report or '',
                report.remarks or '',
            ])

        return self._export_workbook_response(workbook, 'task_reports.xlsx')

    @http.route('/my/hr/admin/payslips', type='http', auth='user', website=True)
    def my_hr_admin_payslips(self, **kwargs):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        fy_value = kwargs.get('fy')
        month_value = kwargs.get('month')

        payroll_period = self._get_selected_payroll_period(fy_value=fy_value, month_value=month_value)
        selected_month = payroll_period['selected_month']
        month_navigation = self._get_month_navigation(selected_month)
        admin_payroll_data = self._get_admin_payroll_data(selected_month)

        return request.render(
            'hr_employee_portal.hr_admin_payslips_page',
            self._prepare_portal_values(None, {
                'admin_payroll_data': admin_payroll_data,
                'month_display': month_navigation.get('month_display', ''),
                'previous_month_value': month_navigation.get('previous_month_value', ''),
                'next_month_value': month_navigation.get('next_month_value', ''),
                'selected_month_value': selected_month.strftime('%Y-%m'),
                'fiscal_year_label': f"FY-{payroll_period['selected_fy_code']}",
                'fiscal_year_code': payroll_period['selected_fy_code'],
                'fiscal_year_range_label': payroll_period['selected_fy_range_label'],
                'fiscal_year_month_options': payroll_period['fiscal_year_month_options'],
                'fiscal_year_options': payroll_period['fiscal_year_options'],
                'selected_fiscal_year_value': payroll_period['selected_fy_value'],
                'selected_fiscal_year_code': payroll_period['selected_fy_code'],
                'payroll_status': (kwargs.get('payroll_status') or '').strip(),
                'payroll_message': (kwargs.get('payroll_message') or '').strip(),
            })
        )

    @http.route('/my/hr/admin/leaves', type='http', auth='user', website=True)
    def my_hr_admin_leaves(self, **kwargs):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        request_state = (kwargs.get('request_state') or '').strip()
        domain = []

        if request_state in ['submitted', 'approved', 'rejected']:
            domain.append(('state', '=', request_state))

        try:
            page = int(kwargs.get('page') or 1)
        except (TypeError, ValueError):
            page = 1

        page = max(page, 1)
        per_page = 10
        offset = (page - 1) * per_page

        request_env = request.env['hr.employee.portal.request'].sudo()
        request_total = request_env.search_count(domain)

        hr_requests = request_env.search(
            domain,
            order='submitted_at desc, id desc',
            limit=per_page,
            offset=offset,
        )

        page_count = max((request_total + per_page - 1) // per_page, 1)
        request_start = offset + 1 if request_total else 0
        request_end = min(offset + per_page, request_total)

        previous_page_url = False
        next_page_url = False

        base_params = {
            'request_state': request_state,
        }

        if page > 1:
            previous_params = dict(base_params)
            previous_params['page'] = page - 1
            previous_page_url = self._build_redirect_url('/my/hr/admin/leaves', previous_params)

        if page < page_count:
            next_params = dict(base_params)
            next_params['page'] = page + 1
            next_page_url = self._build_redirect_url('/my/hr/admin/leaves', next_params)

        return request.render(
            'hr_employee_portal.hr_admin_leaves_page',
            self._prepare_portal_values(None, {
                'hr_requests': hr_requests,
                'selected_request_state': request_state,
                'request_total': request_total,
                'request_page': page,
                'request_page_count': page_count,
                'request_start': request_start,
                'request_end': request_end,
                'previous_page_url': previous_page_url,
                'next_page_url': next_page_url,
                'admin_request_status': (kwargs.get('admin_request_status') or '').strip(),
                'admin_request_message': (kwargs.get('admin_request_message') or '').strip(),
            })
        )
    @http.route('/my/hr/admin/requests/update', type='http', auth='user', methods=['POST'], website=True)
    def my_hr_admin_request_update(self, **post):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        request_id_raw = (post.get('request_id') or '').strip()
        action = (post.get('action') or '').strip()
        admin_remarks = (post.get('admin_remarks') or '').strip()
        request_state = (post.get('request_state') or '').strip()
        page = (post.get('page') or '').strip()

        base_params = {
            'request_state': request_state,
            'page': page,
        }

        try:
            request_id = int(request_id_raw)
        except (TypeError, ValueError):
            return request.redirect(self._build_redirect_url('/my/hr/admin/leaves', {
            'admin_request_status': 'error',
            'admin_request_message': 'Invalid request selected.',
        }))

        portal_request = request.env['hr.employee.portal.request'].sudo().browse(request_id).exists()
        if not portal_request:
            return request.redirect(self._build_redirect_url('/my/hr/admin/leaves', {
                'admin_request_status': 'error',
                'admin_request_message': 'Request not found.',
            }))

        if portal_request.state != 'submitted':
            return request.redirect(self._build_redirect_url('/my/hr/admin/leaves', {
                'admin_request_status': 'error',
                'admin_request_message': 'Only submitted requests can be approved or rejected.',
            }))

        if action not in ['approved', 'rejected']:
            return request.redirect(self._build_redirect_url('/my/hr/admin/leaves', {
                'admin_request_status': 'error',
                'admin_request_message': 'Invalid action selected.',
            }))

        portal_request.write({
            'state': action,
            'admin_remarks': admin_remarks,
            'reviewed_by': request.env.user.id,
            'reviewed_at': fields.Datetime.now(),
        })

        try:
            self._send_employee_request_status_email(portal_request)
        except Exception:
            pass

        base_params.update({
            'admin_request_status': 'success',
            'admin_request_message': f"Request {action} successfully.",
})
        return request.redirect(self._build_redirect_url('/my/hr/admin/leaves', base_params))

    @http.route('/my/hr/admin/documents', type='http', auth='user', website=True)
    def my_hr_admin_documents(self, **kwargs):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        return request.render(
            'hr_employee_portal.hr_admin_documents_page',
            self._prepare_portal_values(None)
        )

    @http.route('/my/hr/admin/payslips', type='http', auth='user', website=True)
    def my_hr_admin_payslips(self, **kwargs):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        fy_value = kwargs.get('fy')
        month_value = kwargs.get('month')

        selected_employee_id = False
        employee_id_raw = (kwargs.get('employee_id') or '').strip()
        if employee_id_raw:
            try:
                selected_employee_id = int(employee_id_raw)
            except (TypeError, ValueError):
                selected_employee_id = False

        employees = request.env['hr.employee'].sudo().search([], order='name asc')

        payroll_period = self._get_selected_payroll_period(fy_value=fy_value, month_value=month_value)
        selected_month = payroll_period['selected_month']
        month_navigation = self._get_month_navigation(selected_month)
        admin_payroll_data = self._get_admin_payroll_data(
            selected_month,
            employee_id=selected_employee_id,
        )

        return request.render(
            'hr_employee_portal.hr_admin_payslips_page',
            self._prepare_portal_values(None, {
                'employees': employees,
                'selected_employee_id': selected_employee_id,
                'admin_payroll_data': admin_payroll_data,
                'month_display': month_navigation.get('month_display', ''),
                'previous_month_value': month_navigation.get('previous_month_value', ''),
                'next_month_value': month_navigation.get('next_month_value', ''),
                'selected_month_value': selected_month.strftime('%Y-%m'),
                'fiscal_year_label': f"FY-{payroll_period['selected_fy_code']}",
                'fiscal_year_code': payroll_period['selected_fy_code'],
                'fiscal_year_range_label': payroll_period['selected_fy_range_label'],
                'fiscal_year_month_options': payroll_period['fiscal_year_month_options'],
                'fiscal_year_options': payroll_period['fiscal_year_options'],
                'selected_fiscal_year_value': payroll_period['selected_fy_value'],
                'selected_fiscal_year_code': payroll_period['selected_fy_code'],
                'payroll_status': (kwargs.get('payroll_status') or '').strip(),
                'payroll_message': (kwargs.get('payroll_message') or '').strip(),
            })
        )
    @http.route('/my/hr/admin/payslips/update', type='http', auth='user', methods=['POST'], website=True)
    def my_hr_admin_payslips_update(self, **post):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        if 'hr.payroll.register.line' not in request.env:
            return request.redirect(self._build_redirect_url('/my/hr/admin/payslips', {
                'payroll_status': 'error',
                'payroll_message': 'Payroll register model is not available.',
        }))

        fy_value = (post.get('fy') or '').strip()
        month_value = (post.get('month') or '').strip()
        employee_id_value = (post.get('employee_id') or '').strip()

        base_params = {
            'fy': fy_value,
            'month': month_value,
            'employee_id': employee_id_value,
        }

        record_id_raw = (post.get('record_id') or '').strip()
        try:
            record_id = int(record_id_raw)
        except (TypeError, ValueError):
            base_params.update({
                'payroll_status': 'error',
                'payroll_message': 'Invalid payroll record selected.',
        })
            return request.redirect(self._build_redirect_url('/my/hr/admin/payslips', base_params))

        payroll_line = request.env['hr.payroll.register.line'].sudo().browse(record_id).exists()
        if not payroll_line:
            base_params.update({
                'payroll_status': 'error',
                'payroll_message': 'Payroll record was not found.',
        })
            return request.redirect(self._build_redirect_url('/my/hr/admin/payslips', base_params))

        payment_date = self._parse_portal_date(post.get('payment_date'))

        project_salary_allowance = self._parse_portal_float(post.get('project_salary'))

        vals = {
            'payment_method': (post.get('payment_method') or '').strip(),
            'basic_salary': self._parse_portal_float(post.get('basic_salary')),
            'basic_actual': self._parse_portal_float(post.get('basic_actual')),

            # Combined portal field: Project Salary / Allowance.
            # We store the edited combined value in project_salary and clear old allowance
            # to avoid double-counting after recalculation.
            'project_salary': project_salary_allowance,
            'allowance': 0.0,

            'project_detail': (post.get('project_detail') or '').strip(),
            'bonus': self._parse_portal_float(post.get('bonus')),
            'bonus_detail': (post.get('bonus_detail') or '').strip(),
            'overtime': self._parse_portal_float(post.get('overtime')),
            'overtime_detail': (post.get('overtime_detail') or '').strip(),
            'income_tax_deduction': self._parse_portal_float(post.get('income_tax_deduction')),
            'other_deductions': self._parse_portal_float(post.get('other_deductions')),
            'deduction_detail': (post.get('deduction_detail') or '').strip(),
            'comments': (post.get('comments') or '').strip(),
            'source': 'manual',
    }

        if payment_date:
            vals['payment_date'] = payment_date

        try:
            payroll_line.write(vals)
            payroll_line.action_recalculate_manual_payroll()
        except ValidationError as error:
            base_params.update({
                'payroll_status': 'error',
                'payroll_message': str(error),
        })
            return request.redirect(self._build_redirect_url('/my/hr/admin/payslips', base_params))
        except Exception:
            base_params.update({
                'payroll_status': 'error',
                'payroll_message': 'Payroll record could not be updated.',
        })
            return request.redirect(self._build_redirect_url('/my/hr/admin/payslips', base_params))

        base_params.update({
            'payroll_status': 'success',
            'payroll_message': 'Payroll record updated and recalculated successfully.',
    })
        return request.redirect(self._build_redirect_url('/my/hr/admin/payslips', base_params))

    @http.route('/my/hr/admin/payslips/export', type='http', auth='user', website=True)
    def my_hr_admin_payslips_export(self, **kwargs):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        fy_value = kwargs.get('fy')
        month_value = kwargs.get('month')
        selected_employee_id = False
        employee_id_raw = (kwargs.get('employee_id') or '').strip()
        if employee_id_raw:
            try:
                selected_employee_id = int(employee_id_raw)
            except (TypeError, ValueError):
                selected_employee_id = False

        payroll_period = self._get_selected_payroll_period(
            fy_value=fy_value,
            month_value=month_value,
        )
        selected_month = payroll_period['selected_month']

        admin_payroll_data = self._get_admin_payroll_data(
            selected_month,
            employee_id=selected_employee_id,
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Payroll'

        headers = [
            '#',
            'Employee ID',
            'Employee Name',
            'Bank',
            'Designation',
            'Payment Method',
            'Basic Salary',
            'Basic Actual',
            'Medical Allowance',
            'Advertised Salary',
            'Project Salary / Allowance',
            'Project / Allowance Detail',
            'Bonus',
            'Bonus Detail',
            'Overtime',
            'OT Detail',
            'Taxable Income',
            'Yearly Income',
            'Income Tax',
            'Other Deductions',
            'Deductions For',
            'Comments',
            'Total',
            'Total Round',
            'Allowance',
            'Net Salary',
            'Rounded',
            'Payment Date',
    ]
        sheet.append(headers)

        for index, row in enumerate(admin_payroll_data.get('rows', []), start=1):
            sheet.append([
                index,
                row.get('employee_code', ''),
                row.get('employee_name', ''),
                row.get('bank_name', ''),
                row.get('designation', ''),
                row.get('payment_method', ''),
                row.get('basic_salary', 0.0),
                row.get('basic_actual', 0.0),
                row.get('medical_allowance', 0.0),
                row.get('advertised_salary', 0.0),
                row.get('project_salary', 0.0),
                row.get('project_value', ''),
                row.get('bonus', 0.0),
                row.get('for_value', ''),
                row.get('overtime', 0.0),
                row.get('ot_detail', ''),
                row.get('taxable_income', 0.0),
                row.get('yearly_income', 0.0),
                row.get('income_tax_deduction', 0.0),
                row.get('other_deductions', 0.0),
                row.get('deduction_for', ''),
                row.get('comments_raw', ''),
                row.get('total', 0.0),
                row.get('total_round', 0.0),
                row.get('total_allowance', 0.0),
                row.get('hr_cost_including_bonus', 0.0),
                row.get('hr_cost_rounded', 0.0),
                row.get('payment_date', ''),
        ])

        for column_cells in sheet.columns:
            max_length = 12
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = str(cell.value or '')
                max_length = max(max_length, min(len(value) + 2, 45))
                sheet.column_dimensions[column_letter].width = max_length

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        file_name = 'payroll_%s.xlsx' % selected_month.strftime('%Y_%m')

        return request.make_response(
            output.getvalue(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', 'attachment; filename="%s"' % file_name),
        ]
    )
    @http.route('/my/hr/admin/payslips/tax-calculator', type='http', auth='user', website=True)
    def my_hr_admin_payslips_tax_calculator(self, **kwargs):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        fy_value = kwargs.get('fy')
        month_value = kwargs.get('month')

        payroll_period = self._get_selected_payroll_period(fy_value=fy_value, month_value=month_value)
        selected_month = payroll_period['selected_month']
        month_navigation = self._get_month_navigation(selected_month)
        admin_payroll_data = self._get_admin_payroll_data(selected_month)

        return request.render(
            'hr_employee_portal.hr_admin_payslips_tax_calculator_page',
            self._prepare_portal_values(None, {
                'admin_payroll_data': admin_payroll_data,
                'month_display': month_navigation.get('month_display', ''),
                'previous_month_value': month_navigation.get('previous_month_value', ''),
                'next_month_value': month_navigation.get('next_month_value', ''),
                'selected_month_value': selected_month.strftime('%Y-%m'),
                'fiscal_year_label': f"FY-{payroll_period['selected_fy_code']}",
                'fiscal_year_code': payroll_period['selected_fy_code'],
                'fiscal_year_range_label': payroll_period['selected_fy_range_label'],
                'fiscal_year_month_options': payroll_period['fiscal_year_month_options'],
                'fiscal_year_options': payroll_period['fiscal_year_options'],
                'selected_fiscal_year_value': payroll_period['selected_fy_value'],
                'selected_fiscal_year_code': payroll_period['selected_fy_code'],
            })
        )

    @http.route('/my/hr/admin/payslips/tax-rates', type='http', auth='user', website=True)
    def my_hr_admin_payslips_tax_rates(self, **kwargs):
        if not self._is_hr_manager():
            return request.redirect('/my/hr')

        fy_value = kwargs.get('fy')
        month_value = kwargs.get('month')

        payroll_period = self._get_selected_payroll_period(fy_value=fy_value, month_value=month_value)
        selected_month = payroll_period['selected_month']
        month_navigation = self._get_month_navigation(selected_month)
        admin_payroll_data = self._get_admin_payroll_data(selected_month)

        return request.render(
            'hr_employee_portal.hr_admin_payslips_tax_rates_page',
            self._prepare_portal_values(None, {
                'admin_payroll_data': admin_payroll_data,
                'month_display': month_navigation.get('month_display', ''),
                'previous_month_value': month_navigation.get('previous_month_value', ''),
                'next_month_value': month_navigation.get('next_month_value', ''),
                'selected_month_value': selected_month.strftime('%Y-%m'),
                'fiscal_year_label': f"FY-{payroll_period['selected_fy_code']}",
                'fiscal_year_code': payroll_period['selected_fy_code'],
                'fiscal_year_range_label': payroll_period['selected_fy_range_label'],
                'fiscal_year_month_options': payroll_period['fiscal_year_month_options'],
                'fiscal_year_options': payroll_period['fiscal_year_options'],
                'selected_fiscal_year_value': payroll_period['selected_fy_value'],
                'selected_fiscal_year_code': payroll_period['selected_fy_code'],
            })
        )
